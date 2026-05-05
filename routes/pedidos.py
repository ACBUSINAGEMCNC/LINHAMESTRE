from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app, session
from models import db, Pedido, Cliente, UnidadeEntrega, Item, PedidoOrdemServico, OrdemServico, Material, Trabalho, PedidoMaterial, ItemPedidoMaterial, ItemMaterial, ItemComposto, PedidoMontagem, ItemPedidoMontagem
from utils import validate_form_data, parse_json_field, generate_next_code, generate_next_os_code
from datetime import datetime
import logging
from sqlalchemy import and_
from sqlalchemy.orm import joinedload
from openpyxl import load_workbook
from datetime import date
import re
import unicodedata

pedidos = Blueprint('pedidos', __name__)
logger = logging.getLogger(__name__)


def _ensure_item_pedido_material_laser_schema():
    try:
        from migrations.add_laser_fields_item_pedido_material import migrate_postgres, migrate_sqlite
        if db.engine.url.drivername.startswith('postgresql'):
            return migrate_postgres()
        return migrate_sqlite()
    except Exception as e:
        logger.warning(f"Erro ao garantir schema de item_pedido_material para laser: {str(e)}")
        return False


def _parse_date_cell(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(v, fmt).date()
            except Exception:
                pass
    return None


def _normalize_header(value):
    if value is None:
        return ''
    s = str(value).strip().lower()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r'[^a-z0-9]+', '_', s)
    s = re.sub(r'_+', '_', s).strip('_')
    return s


def _canonical_column_name(header: str) -> str:
    """Mapeia variações comuns de cabeçalho para o nome canônico usado no importador."""
    h = _normalize_header(header)
    aliases = {
        'cliente': {'cliente', 'nome_cliente'},
        'unidade': {'unidade', 'unidade_entrega', 'unidade_de_entrega', 'unidadeentrega'},
        'item': {'item', 'codigo_acb', 'codigo', 'cod_item', 'cod'},
        'quantidade': {'quantidade', 'qtd', 'qtde'},
        'numero_pedido_cliente': {
            'numero_pedido_cliente', 'n_pedido_cliente', 'numero_pedido', 'n_pedido', 'pedido_cliente', 'oc', 'ordem_compra'
        },
        'prazo_entrega': {'prazo_entrega', 'prazo', 'entrega', 'data_entrega', 'previsao_entrega', 'previsao'},
    }
    for canonical, opts in aliases.items():
        if h in opts:
            return canonical
    return h


def _generate_next_pedido_code():
    ultimo_pedido = Pedido.query.filter(
        (Pedido.numero_pedido != None) & (~Pedido.numero_pedido.like('AUTO-%'))
    ).order_by(Pedido.numero_pedido.desc()).first()
    if ultimo_pedido and ultimo_pedido.numero_pedido:
        try:
            ultimo_numero = int(ultimo_pedido.numero_pedido.split('-')[-1])
            return f"PED-{str(ultimo_numero + 1).zfill(5)}"
        except Exception:
            return "PED-00001"
    return "PED-00001"


@pedidos.route('/pedidos/importar-excel', methods=['GET', 'POST'])
def importar_pedidos_excel():
    if request.method == 'GET':
        session.pop('import_pedidos_excel_rows', None)
        return render_template('pedidos/importar_excel.html')

    arquivo = request.files.get('arquivo')
    if not arquivo or not arquivo.filename:
        flash('Selecione um arquivo XLSX para importar.', 'danger')
        return redirect(url_for('pedidos.importar_pedidos_excel'))

    try:
        wb = load_workbook(arquivo, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f'Erro ao ler XLSX: {e}', 'danger')
        return redirect(url_for('pedidos.importar_pedidos_excel'))

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header_row = row
        break
    if not header_row:
        flash('XLSX vazio ou sem cabeçalho.', 'danger')
        return redirect(url_for('pedidos.importar_pedidos_excel'))

    headers = [_canonical_column_name(h) for h in header_row]
    col_index = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        # Não sobrescrever primeira ocorrência (evita confusão se existir coluna repetida)
        col_index.setdefault(h, i)

    required = ['cliente', 'unidade', 'item', 'quantidade', 'numero_pedido_cliente']
    missing = [c for c in required if c not in col_index]
    if missing:
        encontrados = [h for h in headers if h]
        flash('Colunas obrigatórias ausentes no Excel: ' + ', '.join(missing), 'danger')
        flash('Cabeçalhos encontrados: ' + ', '.join(encontrados[:30]) + ('...' if len(encontrados) > 30 else ''), 'warning')
        return redirect(url_for('pedidos.importar_pedidos_excel'))

    preview_rows = []
    ok_rows = []
    ok_count = 0
    err_count = 0
    warn_count = 0

    prazo_col = col_index.get('prazo_entrega')

    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cliente_nome = (values[col_index['cliente']] or '')
        unidade_nome = (values[col_index['unidade']] or '')
        item_val = (values[col_index['item']] or '')
        quantidade_val = values[col_index['quantidade']]
        numero_pedido_cliente = (values[col_index['numero_pedido_cliente']] or '')
        prazo_entrega_val = values[prazo_col] if prazo_col is not None else None

        errors = []
        warnings = []
        cliente_nome_s = str(cliente_nome).strip()
        unidade_nome_s = str(unidade_nome).strip()
        item_s = str(item_val).strip()
        numero_pedido_cliente_s = str(numero_pedido_cliente).strip()
        prazo = _parse_date_cell(prazo_entrega_val)

        # Pular linhas totalmente vazias
        if not any([cliente_nome_s, unidade_nome_s, item_s, numero_pedido_cliente_s, str(quantidade_val or '').strip(), prazo_entrega_val]):
            continue

        if not cliente_nome_s:
            errors.append('Cliente vazio')
        if not unidade_nome_s:
            errors.append('Unidade vazia')
        if not item_s:
            errors.append('Item vazio')
        if not numero_pedido_cliente_s:
            errors.append('Nº pedido cliente vazio')
        # Prazo de entrega é opcional (pode ficar em branco)
        if prazo_entrega_val not in (None, '') and prazo is None:
            errors.append('Prazo entrega inválido')

        try:
            if isinstance(quantidade_val, float) and quantidade_val.is_integer():
                quantidade_val = int(quantidade_val)
            quantidade_int = int(str(quantidade_val).strip())
            if quantidade_int <= 0:
                errors.append('Quantidade deve ser > 0')
        except Exception:
            quantidade_int = None
            errors.append('Quantidade inválida')

        cliente = None
        unidade = None
        item = None

        if cliente_nome_s:
            cliente = Cliente.query.filter(Cliente.nome.ilike(cliente_nome_s)).first()
            if not cliente:
                errors.append('Cliente não encontrado')

        if cliente and unidade_nome_s:
            unidade = UnidadeEntrega.query.filter(
                UnidadeEntrega.cliente_id == cliente.id,
                UnidadeEntrega.nome.ilike(unidade_nome_s)
            ).first()
            if not unidade:
                errors.append('Unidade não encontrada para o cliente')

        if item_s:
            # Itens/códigos podem vir como número no Excel (ex.: 123 vira 123.0)
            if item_s.endswith('.0'):
                try:
                    item_s = str(int(float(item_s)))
                except Exception:
                    pass
            item = Item.query.filter(Item.codigo_acb.ilike(item_s)).first()
            if not item:
                item = Item.query.filter(Item.nome.ilike(item_s)).first()
            # Item pode não existir no cadastro: nesse caso, vamos importar como nome_item
            if not item:
                warnings.append('Item não cadastrado: será importado como texto (cadastre depois)')

        prazo_str = prazo.strftime('%d/%m/%Y') if prazo else ''
        ok = len(errors) == 0
        if ok:
            ok_count += 1
            if warnings:
                warn_count += 1
            ok_rows.append({
                'cliente_id': cliente.id,
                'unidade_entrega_id': unidade.id,
                'item_id': item.id if item else None,
                'nome_item': None if item else item_s,
                'quantidade': quantidade_int,
                'numero_pedido_cliente': numero_pedido_cliente_s,
                'previsao_entrega': prazo.isoformat() if prazo else None,
            })
        else:
            err_count += 1

        preview_rows.append({
            'row_number': row_number,
            'cliente': cliente_nome_s,
            'unidade': unidade_nome_s,
            'item': item_s,
            'quantidade': '' if quantidade_int is None else quantidade_int,
            'numero_pedido_cliente': numero_pedido_cliente_s,
            'prazo_entrega': prazo_str,
            'ok': ok,
            'errors': errors,
            'warnings': warnings,
        })

    session['import_pedidos_excel_rows'] = ok_rows
    return render_template(
        'pedidos/importar_excel.html',
        preview_rows=preview_rows,
        preview_ok_count=ok_count,
        preview_error_count=err_count,
        preview_warning_count=warn_count,
    )


@pedidos.route('/pedidos/importar-excel/confirmar', methods=['POST'])
def confirmar_importacao_pedidos_excel():
    ok_rows = session.get('import_pedidos_excel_rows')
    if not ok_rows:
        flash('Nenhuma importação pendente. Faça o upload do Excel novamente.', 'warning')
        return redirect(url_for('pedidos.importar_pedidos_excel'))

    try:
        data_entrada = datetime.now().date()

        grupos = {}
        for r in ok_rows:
            k = (r['cliente_id'], r['unidade_entrega_id'], r['numero_pedido_cliente'], r.get('previsao_entrega'))
            grupos.setdefault(k, []).append(r)

        for _, linhas in grupos.items():
            numero_interno = _generate_next_pedido_code()
            for r in linhas:
                previsao_entrega = None
                if r.get('previsao_entrega'):
                    previsao_entrega = datetime.strptime(r['previsao_entrega'], '%Y-%m-%d').date()

                novo_pedido = Pedido(
                    numero_pedido=numero_interno,
                    numero_pedido_cliente=r['numero_pedido_cliente'],
                    cliente_id=r['cliente_id'],
                    unidade_entrega_id=r['unidade_entrega_id'],
                    item_id=r.get('item_id'),
                    nome_item=r.get('nome_item'),
                    quantidade=r['quantidade'],
                    data_entrada=data_entrada,
                    previsao_entrega=previsao_entrega,
                    descricao=None,
                )
                db.session.add(novo_pedido)

        db.session.commit()
        session.pop('import_pedidos_excel_rows', None)
        flash(f'Importação concluída com sucesso! Itens importados: {len(ok_rows)}', 'success')
        return redirect(url_for('pedidos.listar_pedidos'))
    except Exception as e:
        db.session.rollback()
@pedidos.route('/pedidos/gerar-os-multipla', methods=['POST'])
def gerar_ordem_servico_multipla():
    """Rota para gerar ordens de serviço para vários pedidos selecionados"""
    pedidos_ids = request.form.getlist('pedidos[]')
    if not pedidos_ids:
        flash('Selecione pelo menos um pedido para gerar ordens de serviço', 'danger')
        return redirect(url_for('pedidos.listar_pedidos'))

    # Verificar se há pedidos cancelados
    pedidos_cancelados = []
    pedidos_validos = []
    
    for pid in pedidos_ids:
        pedido = Pedido.query.get(pid)
        if not pedido:
            continue
            
        # Verificar se o pedido está cancelado
        if hasattr(pedido, 'cancelado') and pedido.cancelado:
            pedidos_cancelados.append(pedido.id)
        else:
            if pedido.item_id:  # Verificar se tem item cadastrado
                pedidos_validos.append(pedido)
    
    # Se houver pedidos cancelados, informar ao usuário
    if pedidos_cancelados:
        flash(f'Pedidos cancelados foram ignorados (IDs: {", ".join(map(str, pedidos_cancelados))})', 'warning')
    
    # Se não há pedidos válidos, redirecionar de volta
    if not pedidos_validos:
        flash('Não há pedidos válidos para gerar ordens de serviço', 'danger')
        return redirect(url_for('pedidos.listar_pedidos'))
        
    # Agrupar pedidos por item
    grupos = {}
    for pedido in pedidos_validos:
        grupos.setdefault(pedido.item_id, []).append(pedido)
    # Verificar se há múltiplos itens
    if len(grupos) != 1:
        flash('Selecione apenas pedidos do mesmo item para gerar uma Ordem de Serviço', 'warning')
        return redirect(url_for('pedidos.listar_pedidos'))
    pedidos_grupo = list(grupos.values())[0]

    item_principal = Item.query.get(pedidos_grupo[0].item_id)
    if not item_principal:
        flash('Item não encontrado para os pedidos selecionados', 'danger')
        return redirect(url_for('pedidos.listar_pedidos'))
    # Impedir gerar nova OS se já existe
    if any(p.ordens_servico for p in pedidos_grupo):
        # Se for item composto, tenta reconciliar OS existentes (recriar pedidos virtuais por cliente)
        if item_principal.eh_composto:
            ok = reconciliar_os_composto_existente(pedidos_grupo, item_principal)
            if ok:
                try:
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                    ok = False
            if ok:
                flash('OS de item composto já existia: pedidos virtuais foram atualizados por cliente/unidade nas OS dos componentes.', 'success')
            else:
                flash('OS de item composto já existia, mas não foi possível atualizar os pedidos virtuais nas OS dos componentes. Tente novamente ou gere uma OS nova.', 'warning')
            return redirect(url_for('pedidos.listar_pedidos'))

        # ITEM SIMPLES: Recuperar número de OS existente
        os_existente = pedidos_grupo[0].ordens_servico[0].ordem_servico
        existing_num = os_existente.numero

        # Garantir vínculo PedidoOrdemServico para pedidos novos deste grupo
        vinculados = 0
        ja_vinculados = 0
        for p in pedidos_grupo:
            p.numero_oc = existing_num

            ja_associado = False
            for assoc in (p.ordens_servico or []):
                if assoc.ordem_servico_id == os_existente.id:
                    ja_associado = True
                    break
            if ja_associado:
                ja_vinculados += 1
                continue

            assoc_novo = PedidoOrdemServico(
                pedido_id=p.id,
                ordem_servico_id=os_existente.id,
                quantidade_snapshot=p.quantidade,
            )
            db.session.add(assoc_novo)
            vinculados += 1

        db.session.commit()
        if vinculados > 0:
            flash(f'Pedido(s) adicionado(s) na OS {existing_num} com sucesso ({vinculados} novo(s)).', 'success')
        else:
            flash(f'Ordem de Serviço {existing_num} já existe para este item', 'warning')
        return redirect(url_for('pedidos.listar_pedidos'))
    # Verificar se o item é composto
    print(f"🔍 VERIFICANDO TIPO DE ITEM: {item_principal.codigo_acb}")
    print(f"   É composto: {item_principal.eh_composto}")
    
    if item_principal.eh_composto:
        print("   ➡️  SELEÇÃO DE COMPONENTES (ITEM COMPOSTO)")
        quantidade_total_composto = sum((p.quantidade or 0) for p in pedidos_grupo)

        # Componentes elegíveis para OS: apenas tipo_item != 'montagem'
        componentes = []
        for componente_rel in (item_principal.componentes or []):
            item_componente = componente_rel.item_componente
            if not item_componente:
                continue
            if (item_componente.tipo_item or '').strip().lower() == 'montagem':
                continue
            qtd_necessaria = (componente_rel.quantidade or 0) * quantidade_total_composto
            componentes.append({
                'item_id': item_componente.id,
                'codigo_acb': item_componente.codigo_acb,
                'nome': item_componente.nome,
                'tipo_item': (item_componente.tipo_item or 'producao'),
                'quantidade_necessaria': qtd_necessaria,
            })

        # Renderizar tela para o usuário escolher quais componentes gerar OS
        return render_template(
            'pedidos/selecionar_componentes_os.html',
            item_composto=item_principal,
            pedidos_grupo=pedidos_grupo,
            quantidade_total_composto=quantidade_total_composto,
            componentes=componentes,
        )
    else:
        print("   ➡️  GERANDO OS SIMPLES")
        # ITEM SIMPLES: Gerar OS normal
        return gerar_os_item_simples(pedidos_grupo)


@pedidos.route('/pedidos/gerar-os-composto-confirmar', methods=['POST'])
def gerar_os_composto_confirmar():
    """Confirmação: gerar OS apenas para componentes selecionados (ignora montagem)."""
    pedidos_ids = request.form.getlist('pedidos[]')
    item_composto_id = request.form.get('item_composto_id')
    componentes_ids = request.form.getlist('componentes[]')

    if not pedidos_ids or not item_composto_id:
        flash('Dados inválidos para gerar OS do item composto.', 'danger')
        return redirect(url_for('pedidos.listar_pedidos'))

    try:
        item_composto_id_int = int(item_composto_id)
    except Exception:
        flash('Item composto inválido.', 'danger')
        return redirect(url_for('pedidos.listar_pedidos'))

    pedidos_grupo = []
    for pid in pedidos_ids:
        pedido = Pedido.query.get(pid)
        if pedido and not (hasattr(pedido, 'cancelado') and pedido.cancelado) and pedido.item_id:
            pedidos_grupo.append(pedido)

    if not pedidos_grupo:
        flash('Não há pedidos válidos para gerar OS.', 'danger')
        return redirect(url_for('pedidos.listar_pedidos'))

    item_principal = Item.query.get(pedidos_grupo[0].item_id)
    if not item_principal or not item_principal.eh_composto or item_principal.id != item_composto_id_int:
        flash('Item do pedido não corresponde ao item composto informado.', 'danger')
        return redirect(url_for('pedidos.listar_pedidos'))

    # Garantir grupo de um único item
    if any(p.item_id != item_principal.id for p in pedidos_grupo):
        flash('Selecione apenas pedidos do mesmo item para gerar OS.', 'warning')
        return redirect(url_for('pedidos.listar_pedidos'))

    # Componentes selecionados
    try:
        componentes_ids_int = {int(x) for x in componentes_ids}
    except Exception:
        componentes_ids_int = set()

    if not componentes_ids_int:
        flash('Nenhum componente selecionado para gerar OS.', 'warning')
        return redirect(url_for('pedidos.listar_pedidos'))

    return gerar_os_item_composto_seletivo(pedidos_grupo, item_principal, componentes_ids_int)

def gerar_os_item_simples(pedidos_grupo):
    """Gera uma OS normal para item simples"""
    # Gerar número de OS
    numero_os = generate_next_os_code()
    # Criar Ordem de Serviço
    os_nova = OrdemServico(numero=numero_os, data_criacao=datetime.now().date())
    db.session.add(os_nova)
    db.session.flush()
    # Associar pedidos à OS e atualizar campo numero_oc
    for pedido in pedidos_grupo:
        assoc = PedidoOrdemServico(
            pedido_id=pedido.id, 
            ordem_servico_id=os_nova.id,
            quantidade_snapshot=pedido.quantidade
        )
        pedido.numero_oc = numero_os
        db.session.add(assoc)
    db.session.commit()
    flash(f'Ordem de Serviço {numero_os} gerada com sucesso', 'success')
    return redirect(url_for('pedidos.listar_pedidos'))

def gerar_os_item_composto(pedidos_grupo, item_composto):
    """Gera múltiplas OS desmembrando um item composto"""
    try:
        print(f"🔄 INICIANDO DESMEMBRAMENTO DE ITEM COMPOSTO: {item_composto.codigo_acb}")
        print(f"   Pedidos no grupo: {len(pedidos_grupo)}")
        print(f"   Componentes do item: {len(item_composto.componentes)}")
        
        os_geradas = []
        
        # Calcular quantidade total do item composto
        quantidade_total_composto = sum(pedido.quantidade for pedido in pedidos_grupo)
        print(f"   Quantidade total: {quantidade_total_composto}")
        
        # Para cada componente do item composto
        for componente_rel in item_composto.componentes:
            item_componente = componente_rel.item_componente
            quantidade_componente = componente_rel.quantidade * quantidade_total_composto

            # Componentes de montagem não geram OS (vão para fluxo de materiais/montagem)
            if item_componente and (item_componente.tipo_item or '').strip().lower() == 'montagem':
                continue
            
            print(f"   📦 Processando componente: {item_componente.codigo_acb}")
            print(f"      Quantidade necessária: {quantidade_componente}")
            
            # Gerar número de OS para este componente
            numero_os = generate_next_os_code()
            print(f"      Número OS gerado: {numero_os}")
            
            # Criar Ordem de Serviço para o componente
            os_componente = OrdemServico(
                numero=numero_os, 
                data_criacao=datetime.now().date(),
                status='Entrada'
            )
            # Posicionar no final da lista 'Entrada' no Kanban
            try:
                max_pos = db.session.query(db.func.max(OrdemServico.posicao)).filter_by(status='Entrada').scalar()
                os_componente.posicao = (max_pos or 0) + 1
            except Exception as e:
                print(f"      Aviso: não foi possível calcular posição no Kanban (usando padrão 0). Erro: {e}")
                os_componente.posicao = 0
            db.session.add(os_componente)
            db.session.flush()

            # Criar pedidos virtuais por pedido original (preserva cliente/unidade)
            for pedido_original in pedidos_grupo:
                quantidade_virtual = componente_rel.quantidade * (pedido_original.quantidade or 0)
                if quantidade_virtual <= 0:
                    continue

                pedido_virtual = Pedido(
                    cliente_id=pedido_original.cliente_id,
                    unidade_entrega_id=pedido_original.unidade_entrega_id,
                    item_id=item_componente.id,
                    nome_item=f"{item_componente.nome} (Componente de {item_composto.nome})",
                    descricao=f"Componente gerado automaticamente do item composto {item_composto.codigo_acb}",
                    quantidade=quantidade_virtual,
                    data_entrada=datetime.now().date(),
                    numero_pedido=f"AUTO-{numero_os}-{pedido_original.id}",
                    previsao_entrega=pedido_original.previsao_entrega,
                    numero_oc=numero_os
                )
                db.session.add(pedido_virtual)
                db.session.flush()

                assoc = PedidoOrdemServico(
                    pedido_id=pedido_virtual.id,
                    ordem_servico_id=os_componente.id,
                    quantidade_snapshot=pedido_virtual.quantidade
                )
                db.session.add(assoc)
            
            os_geradas.append({
                'numero': numero_os,
                'componente': item_componente.nome,
                'quantidade': quantidade_componente
            })
        
        # Atualizar pedidos originais com referência às OS geradas
        numeros_os = [os['numero'] for os in os_geradas]
        # Resumo curto para caber no campo numero_oc (20 chars).
        # Ex.: "OS-2025-09-018" ou "OS-2025-09-018 (+2)"
        resumo = None
        if numeros_os:
            primeiro = numeros_os[0]
            extra = len(numeros_os) - 1
            resumo = primeiro if extra <= 0 else f"{primeiro} (+{extra})"
        for pedido in pedidos_grupo:
            if resumo:
                pedido.numero_oc = resumo[:20]
        
        db.session.commit()
        print(f"✅ DESMEMBRAMENTO CONCLUÍDO: {len(os_geradas)} OS geradas")
        
        # Mensagem de sucesso detalhada
        detalhes = []
        for os_info in os_geradas:
            detalhes.append(f"OS {os_info['numero']}: {os_info['componente']} (Qtd: {os_info['quantidade']})")
        
        flash(f'Item composto desmembrado com sucesso! Geradas {len(os_geradas)} OS: {"; ".join(detalhes)}', 'success')
        return redirect(url_for('pedidos.listar_pedidos'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao desmembrar item composto: {str(e)}', 'danger')
        return redirect(url_for('pedidos.listar_pedidos'))


def gerar_os_item_composto_seletivo(pedidos_grupo, item_composto, componentes_ids_selecionados):
    """Gera OS apenas para componentes selecionados (ignorando montagem)."""
    try:
        os_geradas = []
        quantidade_total_composto = sum((pedido.quantidade or 0) for pedido in pedidos_grupo)

        for componente_rel in (item_composto.componentes or []):
            item_componente = componente_rel.item_componente
            if not item_componente:
                continue

            if item_componente.id not in componentes_ids_selecionados:
                continue

            if (item_componente.tipo_item or '').strip().lower() == 'montagem':
                continue

            quantidade_componente = (componente_rel.quantidade or 0) * quantidade_total_composto
            if quantidade_componente <= 0:
                continue

            numero_os = generate_next_os_code()

            os_componente = OrdemServico(
                numero=numero_os,
                data_criacao=datetime.now().date(),
                status='Entrada',
            )
            try:
                max_pos = db.session.query(db.func.max(OrdemServico.posicao)).filter_by(status='Entrada').scalar()
                os_componente.posicao = (max_pos or 0) + 1
            except Exception:
                os_componente.posicao = 0

            db.session.add(os_componente)
            db.session.flush()

            # Criar pedidos virtuais por pedido original (preserva cliente/unidade)
            for pedido_original in pedidos_grupo:
                quantidade_virtual = (componente_rel.quantidade or 0) * (pedido_original.quantidade or 0)
                if quantidade_virtual <= 0:
                    continue

                pedido_virtual = Pedido(
                    cliente_id=pedido_original.cliente_id,
                    unidade_entrega_id=pedido_original.unidade_entrega_id,
                    item_id=item_componente.id,
                    nome_item=f"{item_componente.nome} (Componente de {item_composto.nome})",
                    descricao=f"Componente gerado automaticamente do item composto {item_composto.codigo_acb}",
                    quantidade=quantidade_virtual,
                    data_entrada=datetime.now().date(),
                    numero_pedido=f"AUTO-{numero_os}-{pedido_original.id}",
                    previsao_entrega=pedido_original.previsao_entrega,
                    numero_oc=numero_os,
                )
                db.session.add(pedido_virtual)
                db.session.flush()

                assoc = PedidoOrdemServico(
                    pedido_id=pedido_virtual.id,
                    ordem_servico_id=os_componente.id,
                    quantidade_snapshot=pedido_virtual.quantidade,
                )
                db.session.add(assoc)

            os_geradas.append({
                'numero': numero_os,
                'componente': item_componente.nome,
                'quantidade': quantidade_componente,
            })

        # Atualizar pedidos originais com referência às OS geradas
        numeros_os = [os['numero'] for os in os_geradas]
        resumo = None
        if numeros_os:
            primeiro = numeros_os[0]
            extra = len(numeros_os) - 1
            resumo = primeiro if extra <= 0 else f"{primeiro} (+{extra})"

        for pedido in pedidos_grupo:
            if resumo:
                pedido.numero_oc = resumo[:20]

        db.session.commit()

        if os_geradas:
            detalhes = []
            for os_info in os_geradas:
                detalhes.append(f"OS {os_info['numero']}: {os_info['componente']} (Qtd: {os_info['quantidade']})")
            flash(f'Geradas {len(os_geradas)} OS do item composto: {"; ".join(detalhes)}', 'success')
        else:
            flash('Nenhuma OS foi gerada para o item composto.', 'warning')

        return redirect(url_for('pedidos.listar_pedidos'))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao gerar OS do item composto: {str(e)}', 'danger')
        return redirect(url_for('pedidos.listar_pedidos'))


def reconciliar_os_composto_existente(pedidos_grupo, item_composto):
    """Atualiza OS já existentes de item composto, recriando pedidos virtuais por pedido original."""
    try:
        componentes = item_composto.componentes
        if not componentes:
            return False

        atualizou_algum = False

        for componente_rel in componentes:
            item_componente = componente_rel.item_componente
            if not item_componente:
                continue

            # localizar OS existente deste componente via associação de pedidos AUTO
            assoc = (
                db.session.query(PedidoOrdemServico)
                .join(Pedido, PedidoOrdemServico.pedido_id == Pedido.id)
                .filter(
                    Pedido.item_id == item_componente.id,
                    Pedido.numero_pedido.isnot(None),
                    Pedido.numero_pedido.like('AUTO-%'),
                    and_(Pedido.descricao.isnot(None), Pedido.descricao.like('Componente gerado automaticamente%')),
                    Pedido.descricao.like(f"%{item_composto.codigo_acb}%"),
                )
                .order_by(PedidoOrdemServico.id.desc())
                .first()
            )
            if not assoc:
                continue

            os_existente = OrdemServico.query.get(assoc.ordem_servico_id)
            if not os_existente:
                continue

            recriar_pedidos_virtuais_os_componente(os_existente, pedidos_grupo, item_composto, componente_rel)
            atualizou_algum = True

        return atualizou_algum
    except Exception:
        db.session.rollback()
        logger.exception('Erro ao reconciliar OS existente de item composto')
        return False


def recriar_pedidos_virtuais_os_componente(os_componente, pedidos_grupo, item_composto, componente_rel):
    """Remove pedidos virtuais AUTO existentes na OS do componente e recria 1 por pedido original."""
    # Remover associações/pedidos virtuais antigos desta OS
    assocs = list(os_componente.pedidos)
    for pedido_os in assocs:
        pedido = pedido_os.pedido
        if not pedido:
            continue
        if pedido.numero_pedido and pedido.numero_pedido.startswith('AUTO-') and pedido.descricao and pedido.descricao.startswith('Componente gerado automaticamente'):
            db.session.delete(pedido_os)
            db.session.delete(pedido)

    # Recriar pedidos virtuais por pedido original
    for pedido_original in pedidos_grupo:
        quantidade_virtual = componente_rel.quantidade * (pedido_original.quantidade or 0)
        if quantidade_virtual <= 0:
            continue

        pedido_virtual = Pedido(
            cliente_id=pedido_original.cliente_id,
            unidade_entrega_id=pedido_original.unidade_entrega_id,
            item_id=componente_rel.item_componente_id,
            nome_item=f"{componente_rel.item_componente.nome} (Componente de {item_composto.nome})",
            descricao=f"Componente gerado automaticamente do item composto {item_composto.codigo_acb}",
            quantidade=quantidade_virtual,
            data_entrada=datetime.now().date(),
            numero_pedido=f"AUTO-{os_componente.numero}-{pedido_original.id}",
            previsao_entrega=pedido_original.previsao_entrega,
            numero_oc=os_componente.numero
        )
        db.session.add(pedido_virtual)
        db.session.flush()

        assoc = PedidoOrdemServico(
            pedido_id=pedido_virtual.id,
            ordem_servico_id=os_componente.id,
            quantidade_snapshot=pedido_virtual.quantidade
        )
        db.session.add(assoc)

@pedidos.route('/pedidos/gerar-pedido-material-multiplo', methods=['POST'])
def gerar_pedido_material_multiplo():
    """Rota para gerar pedido de material a partir de pedidos selecionados"""
    _ensure_item_pedido_material_laser_schema()
    try:
        logger.info("Iniciando geração de pedido de material múltiplo")
        current_app.logger.info("Rota pedidos/gerar-pedido-material-multiplo recebida")
        
        # Obter IDs dos pedidos selecionados
        form_data = request.form
        current_app.logger.info(f"Formulário completo: {form_data}")
        
        pedidos_ids = request.form.getlist('pedidos[]')
        current_app.logger.info(f"Pedidos IDs recebidos: {pedidos_ids}")
        logger.debug("Pedidos IDs recebidos: %s", pedidos_ids)
        
        # Verificar especificamente os pedidos 4 e 5
        if '4' in pedidos_ids or '5' in pedidos_ids:
            logger.info("*** DETECTADOS PEDIDOS 4 OU 5 NA SELEÇÃO ***")
            # Consultar diretamente no banco de dados
            try:
                conn = db.engine.connect()
                if '4' in pedidos_ids:
                    result = conn.execute(db.text("SELECT * FROM pedido WHERE id = 4"))
                    row = result.fetchone()
                    logger.debug("Pedido 4 no banco: %s", row)
                if '5' in pedidos_ids:
                    result = conn.execute(db.text("SELECT * FROM pedido WHERE id = 5"))
                    row = result.fetchone()
                    logger.debug("Pedido 5 no banco: %s", row)
                conn.close()
            except Exception as e:
                logger.exception("Erro ao consultar pedidos 4/5 diretamente")
            logger.info("*** FIM DA VERIFICAÇÃO ESPECIAL ***")
    except Exception as e:
        logger.exception("Erro ao iniciar o processo de geração de pedido de material")
        flash(f'Erro ao iniciar o processo: {str(e)}', 'danger')
        return redirect(url_for('pedidos.listar_pedidos'))
    
    if not pedidos_ids:
        flash('Selecione pelo menos um pedido para gerar pedido de material', 'danger')
        return redirect(url_for('pedidos.listar_pedidos'))
    
    # Verificar se há pedidos cancelados
    pedidos_cancelados = []
    pedidos_validos = []
    
    for pid in pedidos_ids:
        pedido = Pedido.query.get(pid)
        if not pedido:
            continue
            
        # Verificar se o pedido está cancelado
        if hasattr(pedido, 'cancelado') and pedido.cancelado:
            pedidos_cancelados.append(pedido.id)
        else:
            pedidos_validos.append(pedido)
    
    # Se houver pedidos cancelados, informar ao usuário
    if pedidos_cancelados:
        flash(f'Pedidos cancelados foram ignorados (IDs: {", ".join(map(str, pedidos_cancelados))})', 'warning')
    
    # Se não há pedidos válidos, redirecionar de volta
    if not pedidos_validos:
        flash('Não há pedidos válidos para gerar pedido de material', 'danger')
        return redirect(url_for('pedidos.listar_pedidos'))
    
    # Verificar se todos os pedidos válidos têm itens válidos
    logger.debug("Verificando pedidos válidos...")
    pedidos_sem_item = []
    pedidos_para_processar = []
    
    for pedido in pedidos_validos:
        if not pedido.item_id:
            pedidos_sem_item.append(pedido.id)
        else:
            pedidos_para_processar.append(pedido)
    
    if pedidos_sem_item:
        flash(f'Pedidos sem item associado foram ignorados (IDs: {", ".join(map(str, pedidos_sem_item))})', 'warning')
    
    if not pedidos_para_processar:
        flash('Não há pedidos com itens válidos para gerar pedido de material', 'danger')
        return redirect(url_for('pedidos.listar_pedidos'))
    
    # Agregar materiais de todos os pedidos válidos (incluindo desmembramento de itens compostos)
    # e separar itens de montagem em um Pedido de Montagem
    materiais_agrupados = {}
    materiais_agrupados_qtd = {}
    materiais_laser_agrupados = {}
    itens_montagem_agrupados = {}
    for pedido in pedidos_para_processar:
        item = Item.query.get(pedido.item_id)
        
        if item.eh_composto:
            # ITEM COMPOSTO: Desmembrar e agregar materiais dos componentes
            for componente_rel in item.componentes:
                item_componente = componente_rel.item_componente
                quantidade_componente = componente_rel.quantidade * pedido.quantidade

                if item_componente and (getattr(item_componente, 'tipo_item', None) or 'producao') == 'montagem':
                    itens_montagem_agrupados[item_componente.id] = itens_montagem_agrupados.get(item_componente.id, 0) + quantidade_componente
                    continue
                
                # Buscar materiais do componente
                item_materiais = ItemMaterial.query.filter_by(item_id=item_componente.id).all()
                if not item_materiais and (item_componente.tipo_bruto or '').strip().upper() == 'LASER':
                    descricao_laser = ' '.join(filter(None, [
                        (item_componente.material_laser or '').strip(),
                        (item_componente.espessura_laser or '').strip()
                    ])).strip() or f"Item laser {item_componente.nome}"
                    chave_laser = (item_componente.id, descricao_laser)
                    materiais_laser_agrupados[chave_laser] = materiais_laser_agrupados.get(chave_laser, 0) + quantidade_componente
                    continue
                for item_material in item_materiais:
                    material = Material.query.get(item_material.material_id)
                    if material and material.especifico:
                        qtd_necessaria = (item_material.quantidade or 1) * quantidade_componente
                        if item_material.material_id in materiais_agrupados_qtd:
                            materiais_agrupados_qtd[item_material.material_id] += qtd_necessaria
                        else:
                            materiais_agrupados_qtd[item_material.material_id] = qtd_necessaria
                    else:
                        comprimento_necessario = (item_material.comprimento or 0) * quantidade_componente
                        if item_material.material_id in materiais_agrupados:
                            materiais_agrupados[item_material.material_id] += comprimento_necessario
                        else:
                            materiais_agrupados[item_material.material_id] = comprimento_necessario
        else:
            # ITEM SIMPLES: Processar normalmente
            if item and (getattr(item, 'tipo_item', None) or 'producao') == 'montagem':
                itens_montagem_agrupados[item.id] = itens_montagem_agrupados.get(item.id, 0) + (pedido.quantidade or 0)
                continue
            item_materiais = ItemMaterial.query.filter_by(item_id=pedido.item_id).all()
            if item and not item_materiais and (item.tipo_bruto or '').strip().upper() == 'LASER':
                descricao_laser = ' '.join(filter(None, [
                    (item.material_laser or '').strip(),
                    (item.espessura_laser or '').strip()
                ])).strip() or f"Item laser {item.nome}"
                chave_laser = (item.id, descricao_laser)
                materiais_laser_agrupados[chave_laser] = materiais_laser_agrupados.get(chave_laser, 0) + (pedido.quantidade or 0)
                continue
            for item_material in item_materiais:
                material = Material.query.get(item_material.material_id)
                if material and material.especifico:
                    qtd_necessaria = (item_material.quantidade or 1) * pedido.quantidade
                    if item_material.material_id in materiais_agrupados_qtd:
                        materiais_agrupados_qtd[item_material.material_id] += qtd_necessaria
                    else:
                        materiais_agrupados_qtd[item_material.material_id] = qtd_necessaria
                else:
                    comprimento_necessario = (item_material.comprimento or 0) * pedido.quantidade
                    if item_material.material_id in materiais_agrupados:
                        materiais_agrupados[item_material.material_id] += comprimento_necessario
                    else:
                        materiais_agrupados[item_material.material_id] = comprimento_necessario
    
    if not materiais_agrupados and not materiais_agrupados_qtd and not materiais_laser_agrupados and not itens_montagem_agrupados:
        flash('Nenhum material ou item de montagem associado aos itens dos pedidos selecionados', 'warning')
        return redirect(url_for('pedidos.listar_pedidos'))
    
    codigo_pm = None
    if materiais_agrupados or materiais_agrupados_qtd or materiais_laser_agrupados:
        # Gerar um único código de pedido de material
        codigo_pm = generate_next_code(PedidoMaterial, 'PM', 'numero', padding=5)
        logger.info("Gerando Pedido de Material %s", codigo_pm)

        # Criar um único pedido de material
        pm = PedidoMaterial(
            numero=codigo_pm,
            data_criacao=datetime.now().date()
        )
        db.session.add(pm)
        db.session.flush()
        logger.debug("Pedido de Material %s criado no banco", codigo_pm)

        # Criar associações ItemPedidoMaterial para cada material agrupado
        for material_id, comprimento_total in materiais_agrupados.items():
            assoc = ItemPedidoMaterial(
                pedido_material_id=pm.id,
                material_id=material_id,
                comprimento=comprimento_total
            )
            db.session.add(assoc)

        for material_id, qtd_total in materiais_agrupados_qtd.items():
            assoc = ItemPedidoMaterial(
                pedido_material_id=pm.id,
                material_id=material_id,
                quantidade=qtd_total
            )
            db.session.add(assoc)

        for (item_origem_id, descricao_material), qtd_total in materiais_laser_agrupados.items():
            assoc = ItemPedidoMaterial(
                pedido_material_id=pm.id,
                material_id=None,
                quantidade=int(qtd_total or 0),
                descricao_material=descricao_material,
                item_origem_id=item_origem_id,
                sufixo='LASER'
            )
            db.session.add(assoc)

    codigo_pmont = None
    if itens_montagem_agrupados:
        codigo_pmont = generate_next_code(PedidoMontagem, 'PMT', 'numero', padding=5)
        logger.info("Gerando Pedido de Montagem %s", codigo_pmont)

        pmont = PedidoMontagem(
            numero=codigo_pmont,
            data_criacao=datetime.now().date()
        )
        db.session.add(pmont)
        db.session.flush()

        for item_id, qtd_total in itens_montagem_agrupados.items():
            assoc = ItemPedidoMontagem(
                pedido_montagem_id=pmont.id,
                item_id=item_id,
                quantidade=int(qtd_total or 0)
            )
            db.session.add(assoc)

    # Atualizar numeros para todos os pedidos processados
    for pedido in pedidos_para_processar:
        if codigo_pm:
            pedido.numero_pedido_material = codigo_pm
        if codigo_pmont:
            pedido.numero_pedido_montagem = codigo_pmont

    db.session.commit()

    msg = []
    if codigo_pm:
        msg.append(f"Material: {codigo_pm}")
    if codigo_pmont:
        msg.append(f"Montagem: {codigo_pmont}")
    flash('Pedidos gerados com sucesso: ' + ' | '.join(msg), 'success')
    return redirect(url_for('pedidos.listar_pedidos'))

@pedidos.route('/pedidos/novo', methods=['GET', 'POST'])
def novo_pedido():
    """Rota para cadastrar um novo pedido"""
    if request.method == 'POST':
        cliente_id = request.form.get('cliente_id')
        unidade_entrega_id = request.form.get('unidade_entrega_id')
        if not unidade_entrega_id:
            flash('Unidade de entrega não selecionada!', 'danger')
            return redirect(url_for('pedidos.novo_pedido'))
        
        data_entrada = request.form.get('data_entrada')
        
        # Verificar se há itens no formulário
        itens = []
        index = 0
        while True:
            item_id_key = f'itens[{index}][item_id]'
            if item_id_key not in request.form:
                break
            item_id = request.form.get(item_id_key)
            quantidade = request.form.get(f'itens[{index}][quantidade]')
            nome_item = request.form.get(f'itens[{index}][nome_item]', None)
            if item_id and quantidade:
                itens.append({
                    'item_id': item_id if item_id != 'sem_cadastro' else None,
                    'quantidade': int(quantidade),
                    'nome_item': nome_item
                })
            index += 1
        
        if not itens:
            flash('É necessário adicionar pelo menos um item ao pedido!', 'danger')
            return redirect(url_for('pedidos.novo_pedido'))
        
        # Converter datas
        try:
            data_entrada = datetime.strptime(data_entrada, '%Y-%m-%d').date() if data_entrada else None
            previsao_entrega = datetime.strptime(request.form['previsao_entrega'], '%Y-%m-%d').date() if request.form.get('previsao_entrega') else None
        except ValueError:
            flash('Formato de data inválido!', 'danger')
            return redirect(url_for('pedidos.novo_pedido'))
        
        # Obter número do pedido do cliente (digitado no formulário)
        numero_pedido_cliente = request.form.get('numero_pedido', '').strip()
        
        # Gerar número interno do sistema
        ultimo_pedido = Pedido.query.filter(
            (Pedido.numero_pedido != None) & (~Pedido.numero_pedido.like('AUTO-%'))
        ).order_by(Pedido.numero_pedido.desc()).first()
        if ultimo_pedido and ultimo_pedido.numero_pedido:
            try:
                ultimo_numero = int(ultimo_pedido.numero_pedido.split('-')[-1])
                novo_numero = f"PED-{str(ultimo_numero + 1).zfill(5)}"
            except (ValueError, IndexError):
                novo_numero = "PED-00001"
        else:
            novo_numero = "PED-00001"
        
        # Criar pedidos para cada item
        for item in itens:
            novo_pedido = Pedido(
                numero_pedido=novo_numero,
                numero_pedido_cliente=numero_pedido_cliente if numero_pedido_cliente else None,
                cliente_id=cliente_id,
                unidade_entrega_id=unidade_entrega_id,
                item_id=item['item_id'],
                nome_item=item['nome_item'] if not item['item_id'] else None,
                quantidade=item['quantidade'],
                data_entrada=data_entrada,
                previsao_entrega=previsao_entrega,
                descricao=request.form.get('descricao')
            )
            db.session.add(novo_pedido)
        
        db.session.commit()
        msg_pedido = f'Pedido {numero_pedido_cliente}' if numero_pedido_cliente else 'Pedido'
        flash(f'{msg_pedido} criado com sucesso com {len(itens)} item(ns)!', 'success')
        return redirect(url_for('pedidos.listar_pedidos'))
    
    clientes = Cliente.query.all()
    itens = Item.query.all()
    return render_template('pedidos/novo.html', clientes=clientes, itens=itens)

@pedidos.route('/pedidos')
def listar_pedidos():
    """Rota para listar todos os pedidos"""
    # Por padrão, ocultar pedidos entregues e cancelados (melhor visualização)
    # O usuário pode usar filtros para vê-los
    mostrar_todos = request.args.get('mostrar_todos', '0') == '1'
    
    query = Pedido.query.options(
        joinedload(Pedido.item),
        joinedload(Pedido.cliente),
        joinedload(Pedido.unidade_entrega)
    ).filter(
        (Pedido.numero_pedido == None) | (~Pedido.numero_pedido.like('AUTO-%'))
    )
    
    if not mostrar_todos:
        # Ocultar entregues e cancelados por padrão
        query = query.filter(
            (Pedido.cancelado == False) | (Pedido.cancelado == None),
            (Pedido.data_entrega == None)
        )
    
    pedidos = query.all()
    for pedido in pedidos:
        logger.debug("Pedido ID %s: numero_pedido_material = %s", pedido.id, pedido.numero_pedido_material if pedido.numero_pedido_material else 'N/A')
    clientes = Cliente.query.all()
    return render_template('pedidos/listar.html', pedidos=pedidos, clientes=clientes)

@pedidos.route('/pedidos/editar/<int:pedido_id>', methods=['GET', 'POST'])
def editar_pedido(pedido_id):
    """Rota para editar um pedido existente"""
    pedido = Pedido.query.get_or_404(pedido_id)
    
    if request.method == 'POST':
        # Validação de dados
        errors = validate_form_data(request.form, ['cliente_id', 'unidade_entrega_id', 'quantidade', 'data_entrada'])
        if errors:
            for error in errors:
                flash(error, 'danger')
            clientes = Cliente.query.all()
            itens = Item.query.all()
            unidades = UnidadeEntrega.query.filter_by(cliente_id=pedido.cliente_id).all()
            return render_template('pedidos/editar.html', pedido=pedido, clientes=clientes, itens=itens, unidades=unidades)
        
        cliente_id = request.form['cliente_id']
        unidade_entrega_id = request.form['unidade_entrega_id']
        if not unidade_entrega_id:
            flash('Unidade de entrega não selecionada!', 'danger')
            clientes = Cliente.query.all()
            itens = Item.query.all()
            unidades = UnidadeEntrega.query.filter_by(cliente_id=pedido.cliente_id).all()
            return render_template('pedidos/editar.html', pedido=pedido, clientes=clientes, itens=itens, unidades=unidades)
        
        tipo_item = request.form.get('tipo_item')
        
        # Validar quantidade
        try:
            quantidade = int(request.form['quantidade'])
            if quantidade <= 0:
                flash('A quantidade deve ser maior que zero', 'danger')
                clientes = Cliente.query.all()
                itens = Item.query.all()
                unidades = UnidadeEntrega.query.filter_by(cliente_id=pedido.cliente_id).all()
                return render_template('pedidos/editar.html', pedido=pedido, clientes=clientes, itens=itens, unidades=unidades)
        except ValueError:
            flash('A quantidade deve ser um número inteiro', 'danger')
            clientes = Cliente.query.all()
            itens = Item.query.all()
            unidades = UnidadeEntrega.query.filter_by(cliente_id=pedido.cliente_id).all()
            return render_template('pedidos/editar.html', pedido=pedido, clientes=clientes, itens=itens, unidades=unidades)
        
        # Validar data de entrada
        try:
            data_entrada = datetime.strptime(request.form['data_entrada'], '%Y-%m-%d').date()
        except ValueError:
            flash('Data de entrada inválida', 'danger')
            clientes = Cliente.query.all()
            itens = Item.query.all()
            unidades = UnidadeEntrega.query.filter_by(cliente_id=pedido.cliente_id).all()
            return render_template('pedidos/editar.html', pedido=pedido, clientes=clientes, itens=itens, unidades=unidades)
        
        numero_pedido_cliente = request.form.get('numero_pedido_cliente', '').strip()
        
        # Validar previsão de entrega
        previsao_entrega = None
        if 'previsao_entrega' in request.form and request.form['previsao_entrega']:
            try:
                previsao_entrega = datetime.strptime(request.form['previsao_entrega'], '%Y-%m-%d').date()
            except ValueError:
                flash('Data de previsão de entrega inválida', 'danger')
                clientes = Cliente.query.all()
                itens = Item.query.all()
                unidades = UnidadeEntrega.query.filter_by(cliente_id=pedido.cliente_id).all()
                return render_template('pedidos/editar.html', pedido=pedido, clientes=clientes, itens=itens, unidades=unidades)
            
        descricao = request.form.get('descricao', '')
        
        # Verificar se o material foi comprado
        material_comprado = 'material_comprado' in request.form
        
        # Atualizar dados básicos
        pedido.cliente_id = cliente_id
        pedido.unidade_entrega_id = unidade_entrega_id
        pedido.quantidade = quantidade
        pedido.data_entrada = data_entrada
        pedido.numero_pedido_cliente = numero_pedido_cliente if numero_pedido_cliente else None
        pedido.previsao_entrega = previsao_entrega
        pedido.descricao = descricao
        pedido.material_comprado = material_comprado
        
        # Se não estiver associado a uma OS, permitir alterar o item
        if not pedido.numero_oc:
            if tipo_item == 'cadastrado':
                item_id = request.form.get('item_id')
                if item_id:
                    pedido.item_id = item_id
                    pedido.nome_item = None
            else:  # sem_cadastro
                nome_item = request.form.get('nome_item')
                if nome_item:
                    # Verificar se já existe um item com este nome
                    item_existente = Item.query.filter_by(nome=nome_item).first()
                    if item_existente:
                        pedido.item_id = item_existente.id
                        pedido.nome_item = None
                    else:
                        pedido.item_id = None
                        pedido.nome_item = nome_item
        
        db.session.commit()
        flash('Pedido atualizado com sucesso!', 'success')
        return redirect(url_for('pedidos.listar_pedidos'))
    
    clientes = Cliente.query.all()
    itens = Item.query.all()
    unidades = UnidadeEntrega.query.filter_by(cliente_id=pedido.cliente_id).all()
    
    return render_template('pedidos/editar.html', 
                          pedido=pedido, 
                          clientes=clientes, 
                          itens=itens,
                          unidades=unidades)

@pedidos.route('/pedidos/toggle-compra/<int:pedido_id>', methods=['POST'])
def toggle_compra_material(pedido_id):
    """Rota para alternar o status de compra de material de um pedido"""
    pedido = Pedido.query.get_or_404(pedido_id)
    pedido.material_comprado = not pedido.material_comprado
    db.session.commit()
    
    status = "comprado" if pedido.material_comprado else "não comprado"
    flash(f'Material marcado como {status}!', 'success')
    
    # Se a requisição for AJAX, retornar JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({
            'success': True, 
            'material_comprado': pedido.material_comprado,
            'message': f'Material marcado como {status}!'
        })
    
    # Caso contrário, redirecionar para a página de listagem
    return redirect(url_for('pedidos.listar_pedidos'))

@pedidos.route('/pedidos/cancelar/<int:pedido_id>', methods=['POST'])
def cancelar_pedido(pedido_id):
    """Rota para cancelar um pedido (não exclui do banco)"""
    from flask import session
    pedido = Pedido.query.get_or_404(pedido_id)
    logger.info("Tentando cancelar pedido ID %s, status atual: cancelado=%s", pedido_id, pedido.cancelado)
    if pedido.ordens_servico:
        flash('Não é possível cancelar um pedido associado a uma Ordem de Serviço!', 'danger')
        return redirect(url_for('pedidos.listar_pedidos'))
    motivo = request.form.get('motivo_cancelamento')
    if not motivo:
        flash('É necessário informar o motivo do cancelamento!', 'danger')
        return redirect(url_for('pedidos.listar_pedidos'))
    pedido.cancelado = True
    pedido.motivo_cancelamento = motivo
    pedido.cancelado_por = session.get('usuario_nome', 'Desconhecido')
    pedido.data_cancelamento = datetime.now()
    db.session.commit()
    logger.info("Cancelado pedido ID %s com sucesso, novo status: cancelado=%s", pedido_id, pedido.cancelado)
    flash(f'Pedido cancelado com sucesso!', 'success')
    return redirect(url_for('pedidos.listar_pedidos'))

@pedidos.route('/pedidos/cadastrar-item/<int:pedido_id>', methods=['GET', 'POST'])
def cadastrar_item_pedido(pedido_id):
    """Rota para cadastrar um item a partir de um pedido"""
    from routes.itens import novo_item
    
    pedido = Pedido.query.get_or_404(pedido_id)
    
    # Verificar se o pedido já tem um item cadastrado
    if pedido.item_id:
        flash('Este pedido já possui um item cadastrado!', 'warning')
        return redirect(url_for('pedidos.listar_pedidos'))
    
    # Verificar se já existe um item com o mesmo nome
    item_existente = Item.query.filter_by(nome=pedido.nome_item).first()
    
    if item_existente:
        # Atualizar todos os pedidos com o mesmo nome_item
        pedidos_para_atualizar = Pedido.query.filter_by(nome_item=pedido.nome_item, item_id=None).all()
        for p in pedidos_para_atualizar:
            p.item_id = item_existente.id
            p.nome_item = None
        
        db.session.commit()
        flash(f'Item já existente! Todos os pedidos com "{pedido.nome_item}" foram atualizados.', 'success')
        return redirect(url_for('pedidos.listar_pedidos'))
    
    if request.method == 'POST':
        # Redirecionar para a função de novo item com os dados do pedido
        return novo_item()
    
    materiais = Material.query.all()
    trabalhos = Trabalho.query.all()
    return render_template('itens/cadastrar_pedido.html', pedido=pedido, materiais=materiais, trabalhos=trabalhos)
