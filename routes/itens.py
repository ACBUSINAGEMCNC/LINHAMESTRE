from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from werkzeug.utils import secure_filename
import os
import json
import io
import threading
import uuid
from datetime import datetime
from decimal import Decimal, InvalidOperation
import requests
from models import db, Item, Material, Trabalho, ItemMaterial, ItemTrabalho, Pedido, ArquivoCNC, ItemComposto, EstoquePecas
from utils import validate_form_data, save_file, generate_next_code, parse_json_field
from flask import current_app, g
from sqlalchemy import func
from sqlalchemy.orm import selectinload
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

try:
    from PIL import Image
except Exception:
    Image = None

itens = Blueprint('itens', __name__)
ADMIN_MASTER_EMAIL = 'admin@acbusinagem.com.br'
_EXPORTACOES_VALORES_ITENS = {}
_EXPORTACOES_VALORES_ITENS_LOCK = threading.Lock()


def _usuario_pode_ver_valores(usuario=None):
    usuario = usuario or getattr(g, 'usuario', None)
    if not usuario:
        return False
    if (getattr(usuario, 'email', '') or '').strip().lower() == ADMIN_MASTER_EMAIL and getattr(usuario, 'nivel_acesso', None) == 'admin':
        return True
    return bool(getattr(usuario, 'acesso_valores_itens', False))


def _parse_valor_item(raw):
    # Se já for número, retorna diretamente
    if isinstance(raw, (int, float)):
        return float(raw)
    
    txt = str(raw or '').strip()
    if not txt:
        return 0.0
    txt = txt.replace('R$', '').replace('.', '').replace(',', '.').strip()
    try:
        return float(Decimal(txt))
    except (InvalidOperation, ValueError) as e:
        raise ValueError(f'Valor do item inválido: "{raw}" -> "{txt}"')


def _parse_percentual(raw):
    # Se já for número, retorna diretamente
    if isinstance(raw, (int, float)):
        return float(raw)
    
    txt = str(raw or '').strip()
    if not txt:
        return 0.0
    txt = txt.replace('%', '').replace('.', '').replace(',', '.').strip()
    try:
        return float(Decimal(txt))
    except (InvalidOperation, ValueError) as e:
        raise ValueError(f'Percentual de imposto inválido: "{raw}" -> "{txt}"')


def _valor_planilha_preenchido(raw):
    if raw is None:
        return False
    # Se for número, considerar como preenchido inclusive quando for 0
    if isinstance(raw, (int, float)):
        return True
    # Se for string, verificar se não está vazia após strip
    return str(raw).strip() != ''


def _valor_realmente_alterado(raw, valor_atual):
    """Verifica se o valor da planilha é realmente diferente do valor atual"""
    if raw is None:
        return False
    
    # Converter para número para comparação
    try:
        if isinstance(raw, (int, float)):
            valor_planilha = float(raw)
        else:
            # Se for string, fazer parsing
            txt = str(raw or '').strip()
            if not txt:
                return False
            txt = txt.replace('R$', '').replace('.', '').replace(',', '.').strip()
            valor_planilha = float(Decimal(txt))
        
        # Comparar com valor atual (considerando None como 0)
        valor_atual_num = float(valor_atual or 0)
        
        return abs(valor_planilha - valor_atual_num) > 0.001  # Tolerância para diferenças pequenas
        
    except (InvalidOperation, ValueError):
        return False


def _apply_campos_financeiros(item, form_data):
    item.valor_item = _parse_valor_item(form_data.get('valor_item', item.valor_item or 0))
    item.valor_material = _parse_valor_item(form_data.get('valor_material', item.valor_material or 0))
    item.outros_custos = _parse_valor_item(form_data.get('outros_custos', item.outros_custos or 0))
    item.imposto_percentual = _parse_percentual(form_data.get('imposto_percentual', item.imposto_percentual or 0))


def _build_preview_valores_rows(rows):
    return {
        'preview_rows': rows,
        'preview_ok_count': sum(1 for r in rows if r.get('ok')),
        'preview_error_count': sum(1 for r in rows if not r.get('ok')),
    }


def _get_item_imagem_bytes(item):
    file_path = getattr(item, 'imagem', None)
    if not file_path:
        return None, None

    try:
        if file_path.startswith('http://') or file_path.startswith('https://'):
            resp = requests.get(file_path, timeout=(3, 5))
            if resp.status_code != 200:
                return None, None
            return resp.content, file_path

        if file_path.startswith('supabase://') or file_path.startswith('supabase:/'):
            public_url = _build_supabase_public_url_from_file_path(file_path)
            if not public_url:
                return None, None
            resp = requests.get(public_url, timeout=(3, 5))
            if resp.status_code != 200:
                return None, public_url
            return resp.content, public_url

        normalized = file_path.replace('\\', '/').lstrip('/')
        if normalized.startswith('uploads/'):
            normalized = normalized[len('uploads/'):]

        if normalized.startswith('imagens/'):
            filename = normalized.split('/', 1)[1]
            local_path = os.path.join(current_app.config['UPLOAD_FOLDER_IMAGENS'], filename)
            if not os.path.exists(local_path):
                return None, None
            with open(local_path, 'rb') as f:
                return f.read(), getattr(item, 'imagem_path', None)
    except Exception:
        return None, getattr(item, 'imagem_path', None)

    return None, getattr(item, 'imagem_path', None)


def _build_excel_safe_image_stream(image_bytes):
    if not image_bytes:
        return None

    if Image is None:
        return io.BytesIO(image_bytes)

    try:
        source_stream = io.BytesIO(image_bytes)
        with Image.open(source_stream) as img:
            normalized = img.convert('RGB') if img.mode not in ('RGB', 'L') else img.copy()
            output = io.BytesIO()
            normalized.save(output, format='PNG')
            output.seek(0)
            return output
    except Exception:
        return None


def _set_exportacao_valores_status(export_id, **kwargs):
    with _EXPORTACOES_VALORES_ITENS_LOCK:
        exportacao = _EXPORTACOES_VALORES_ITENS.get(export_id)
        if not exportacao:
            return
        exportacao.update(kwargs)


def _carregar_quantidades_estoque_por_item():
    rows = (
        db.session.query(
            EstoquePecas.item_id,
            func.coalesce(func.sum(EstoquePecas.quantidade), 0)
        )
        .group_by(EstoquePecas.item_id)
        .all()
    )
    return {item_id: int(quantidade or 0) for item_id, quantidade in rows}


def _gerar_planilha_valores_itens(progress_callback=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Valores Itens'
    headers = ['item', 'codigo_acb', 'tipo', 'quantidade_estoque', 'valor_item', 'valor_material', 'outros_custos', 'imposto_percentual']
    ws.append(headers)

    header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    currency_fmt = 'R$ #,##0.00'
    percent_fmt = '0.00'

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.freeze_panes = 'B2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    itens_cadastrados = Item.query.order_by(Item.nome.asc()).all()
    quantidades_estoque = _carregar_quantidades_estoque_por_item()
    total_itens = max(len(itens_cadastrados), 1)

    if progress_callback:
        progress_callback(percentual=5, mensagem='Preparando planilha...')

    for index, item in enumerate(itens_cadastrados, start=1):
        tipo_item = 'Composto' if item.eh_composto else ('Montagem' if (item.tipo_item or '') == 'montagem' else 'Produção')
        quantidade_estoque = quantidades_estoque.get(item.id, 0)
        ws.append([
            (item.nome or '').strip().replace('\n', '').replace('\r', ''),
            item.codigo_acb,
            tipo_item,
            quantidade_estoque,
            float(item.valor_item or 0),
            float(item.valor_material or 0),
            float(item.outros_custos or 0),
            float(item.imposto_percentual or 0),
        ])

        current_row = ws.max_row
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center', vertical='center')

        for col in [5, 6, 7]:
            ws.cell(row=current_row, column=col).number_format = currency_fmt
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='right', vertical='center')

        ws.cell(row=current_row, column=8).number_format = percent_fmt
        ws.cell(row=current_row, column=8).alignment = Alignment(horizontal='right', vertical='center')

        if progress_callback:
            percentual_item = 5 + int((index / total_itens) * 85)
            progress_callback(percentual=min(percentual_item, 95), mensagem=f'Processando item {index} de {total_itens}...')

    column_widths = {
        'A': 38,
        'B': 14,
        'C': 16,
        'D': 18,
        'E': 14,
        'F': 16,
        'G': 16,
        'H': 14,
    }
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width

    if progress_callback:
        progress_callback(percentual=97, mensagem='Finalizando arquivo...')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    if progress_callback:
        progress_callback(percentual=100, mensagem='Planilha pronta para download.')

    return output


def _executar_exportacao_valores_job(app, export_id):
    with app.app_context():
        try:
            output = _gerar_planilha_valores_itens(
                progress_callback=lambda percentual, mensagem: _set_exportacao_valores_status(
                    export_id,
                    status='processando',
                    percentual=percentual,
                    mensagem=mensagem,
                )
            )
            _set_exportacao_valores_status(
                export_id,
                status='pronto',
                percentual=100,
                mensagem='Planilha pronta para download.',
                arquivo=output.getvalue(),
                nome_arquivo='valores_itens.xlsx',
            )
        except Exception as exc:
            _set_exportacao_valores_status(
                export_id,
                status='erro',
                mensagem=f'Falha ao gerar planilha: {exc}',
            )


def _require_valores_access():
    if not _usuario_pode_ver_valores():
        flash('Você não tem permissão para acessar os valores dos itens.', 'danger')
        return redirect(url_for('main.index'))
    return None

@itens.route('/itens')
def listar_itens():
    """Rota para listar todos os itens"""
    itens = Item.query.options(
        selectinload(Item.materiais),
        selectinload(Item.trabalhos),
    ).all()
    return render_template('itens/listar.html', itens=itens)


@itens.route('/itens/valores')
def listar_valores_itens():
    acesso_negado = _require_valores_access()
    if acesso_negado:
        return acesso_negado
    itens_cadastrados = Item.query.order_by(Item.nome.asc()).all()
    return render_template('itens/valores.html', itens=itens_cadastrados)


@itens.route('/itens/valores/exportar')
def exportar_planilha_valores_itens():
    acesso_negado = _require_valores_access()
    if acesso_negado:
        return acesso_negado

    if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return redirect(url_for('itens.listar_valores_itens'))

    export_id = uuid.uuid4().hex
    with _EXPORTACOES_VALORES_ITENS_LOCK:
        _EXPORTACOES_VALORES_ITENS[export_id] = {
            'status': 'processando',
            'percentual': 0,
            'mensagem': 'Iniciando exportação...',
            'arquivo': None,
            'nome_arquivo': 'valores_itens.xlsx',
        }

    worker = threading.Thread(
        target=_executar_exportacao_valores_job,
        args=(current_app._get_current_object(), export_id),
        daemon=True,
    )
    worker.start()

    return jsonify({
        'ok': True,
        'export_id': export_id,
        'status_url': url_for('itens.status_exportacao_planilha_valores_itens', export_id=export_id),
        'download_url': url_for('itens.download_exportacao_planilha_valores_itens', export_id=export_id),
    })


@itens.route('/itens/valores/exportar/status/<export_id>')
def status_exportacao_planilha_valores_itens(export_id):
    acesso_negado = _require_valores_access()
    if acesso_negado:
        return acesso_negado

    with _EXPORTACOES_VALORES_ITENS_LOCK:
        exportacao = _EXPORTACOES_VALORES_ITENS.get(export_id)

    if not exportacao:
        return jsonify({'ok': False, 'mensagem': 'Exportação não encontrada.'}), 404

    return jsonify({
        'ok': True,
        'status': exportacao.get('status'),
        'percentual': exportacao.get('percentual', 0),
        'mensagem': exportacao.get('mensagem', ''),
        'pronto': exportacao.get('status') == 'pronto',
        'erro': exportacao.get('status') == 'erro',
    })


@itens.route('/itens/valores/exportar/download/<export_id>')
def download_exportacao_planilha_valores_itens(export_id):
    acesso_negado = _require_valores_access()
    if acesso_negado:
        return acesso_negado

    with _EXPORTACOES_VALORES_ITENS_LOCK:
        exportacao = _EXPORTACOES_VALORES_ITENS.get(export_id)

    if not exportacao:
        flash('Exportação não encontrada.', 'danger')
        return redirect(url_for('itens.listar_valores_itens'))

    if exportacao.get('status') != 'pronto' or not exportacao.get('arquivo'):
        flash('A planilha ainda não está pronta para download.', 'warning')
        return redirect(url_for('itens.listar_valores_itens'))

    output = io.BytesIO(exportacao['arquivo'])
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=exportacao.get('nome_arquivo', 'valores_itens.xlsx'),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@itens.route('/itens/valores/importar', methods=['GET', 'POST'])
def importar_planilha_valores_itens():
    acesso_negado = _require_valores_access()
    if acesso_negado:
        return acesso_negado

    if request.method == 'GET':
        session.pop('import_valores_itens_rows', None)
        return render_template('itens/importar_valores.html')

    arquivo = request.files.get('arquivo')
    if not arquivo or not arquivo.filename:
        flash('Selecione um arquivo XLSX para importar.', 'danger')
        return redirect(url_for('itens.importar_planilha_valores_itens'))

    try:
        wb = load_workbook(arquivo, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f'Erro ao ler XLSX: {e}', 'danger')
        return redirect(url_for('itens.importar_planilha_valores_itens'))

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header_row = row
        break
    if not header_row:
        flash('XLSX vazio ou sem cabeçalho.', 'danger')
        return redirect(url_for('itens.importar_planilha_valores_itens'))

    headers = [str(h or '').strip().lower() for h in header_row]
    col_index = {h: i for i, h in enumerate(headers) if h}
    required_cols = ['codigo_acb', 'item', 'valor_item', 'valor_material', 'outros_custos', 'imposto_percentual']
    if any(col not in col_index for col in required_cols):
        flash('A planilha precisa conter as colunas codigo_acb, item, valor_item, valor_material, outros_custos e imposto_percentual.', 'danger')
        return redirect(url_for('itens.importar_planilha_valores_itens'))

    preview_rows = []
    ok_rows = []
    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        codigo_acb = str(values[col_index['codigo_acb']] or '').strip().replace('\n', '').replace('\r', '')
        nome = str(values[col_index['item']] or '').strip().replace('\n', '').replace('\r', '')
        valor_raw = values[col_index['valor_item']]
        valor_material_raw = values[col_index['valor_material']]
        outros_custos_raw = values[col_index['outros_custos']]
        imposto_percentual_raw = values[col_index['imposto_percentual']]
        
        # Buscar item primeiro para poder comparar valores
        item = Item.query.filter(Item.codigo_acb == codigo_acb).first() if codigo_acb else None
        
        # Verificar se os valores foram preenchidos na planilha (simplificado)
        valor_item_informado = _valor_planilha_preenchido(valor_raw)
        valor_material_informado = _valor_planilha_preenchido(valor_material_raw)
        outros_custos_informado = _valor_planilha_preenchido(outros_custos_raw)
        imposto_percentual_informado = _valor_planilha_preenchido(imposto_percentual_raw)
        algum_campo_informado = any([
            valor_item_informado,
            valor_material_informado,
            outros_custos_informado,
            imposto_percentual_informado,
        ])

        # Sempre incluir itens válidos na preview, mesmo que não tenham alterações
        if not codigo_acb or not item:
            continue

        errors = []
        if not codigo_acb:
            errors.append('Código ACB vazio')
        if not item:
            errors.append(f'Item não encontrado com código ACB: {codigo_acb}')
        if not nome:
            errors.append('Nome do item vazio')
        # Removida a validação que exigia algum campo alterado

        valor_item = None
        if valor_item_informado:
            try:
                valor_item = _parse_valor_item(valor_raw)
                if valor_item < 0:
                    errors.append('Valor do item não pode ser negativo')
            except ValueError as e:
                valor_item = None
                errors.append(str(e))

        valor_material = None
        if valor_material_informado:
            try:
                valor_material = _parse_valor_item(valor_material_raw)
                if valor_material < 0:
                    errors.append('Valor material não pode ser negativo')
            except ValueError as e:
                valor_material = None
                errors.append(str(e))

        outros_custos = None
        if outros_custos_informado:
            try:
                outros_custos = _parse_valor_item(outros_custos_raw)
                if outros_custos < 0:
                    errors.append('Outros custos não podem ser negativos')
            except ValueError as e:
                outros_custos = None
                errors.append(str(e))

        imposto_percentual = None
        if imposto_percentual_informado:
            try:
                imposto_percentual = _parse_percentual(imposto_percentual_raw)
                if imposto_percentual < 0:
                    errors.append('Percentual de imposto não pode ser negativo')
            except ValueError as e:
                imposto_percentual = None
                errors.append(str(e))

        ok = len(errors) == 0
        row_data = {
            'row_number': row_number,
            'codigo_acb': codigo_acb,
            'nome': nome,
            'item_id': item.id if item else None,
            'valor_item': valor_item,
            'valor_material': valor_material,
            'outros_custos': outros_custos,
            'imposto_percentual': imposto_percentual,
            'valor_item_atual': item.valor_item if item else 0,
            'valor_material_atual': item.valor_material if item else 0,
            'outros_custos_atual': item.outros_custos if item else 0,
            'imposto_percentual_atual': item.imposto_percentual if item else 0,
            'valor_item_informado': valor_item_informado,
            'valor_material_informado': valor_material_informado,
            'outros_custos_informado': outros_custos_informado,
            'imposto_percentual_informado': imposto_percentual_informado,
            'ok': ok,
            'errors': errors,
        }
        preview_rows.append(row_data)
        if ok:
            ok_rows.append({
                'row_number': row_number,
                'codigo_acb': codigo_acb,
                'nome': nome,
                'item_id': item.id,
                'valor_item': valor_item,
                'valor_material': valor_material,
                'outros_custos': outros_custos,
                'imposto_percentual': imposto_percentual,
                'valor_item_informado': valor_item_informado,
                'valor_material_informado': valor_material_informado,
                'outros_custos_informado': outros_custos_informado,
                'imposto_percentual_informado': imposto_percentual_informado,
            })

    # Gerar ID único para esta importação
    import uuid
    import_id = str(uuid.uuid4())
    
    # Salvar dados no banco em vez da sessão
    from models import db
    import json
    from sqlalchemy import text
    
    # Criar tabela temporária se não existir
    try:
        db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS import_valores_temp (
            id VARCHAR(36) PRIMARY KEY,
            dados JSONB NOT NULL,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """))
        db.session.commit()
    except:
        pass
    
    # Salvar dados no banco
    db.session.execute(
        text("INSERT INTO import_valores_temp (id, dados) VALUES (:id, :dados)"),
        {"id": import_id, "dados": json.dumps(ok_rows)}
    )
    db.session.commit()
    
    # Renderizar preview com import_id
    return render_template('itens/importar_valores.html', import_id=import_id, **_build_preview_valores_rows(preview_rows))


@itens.route('/itens/valores/importar/confirmar/<import_id>', methods=['POST'])
def confirmar_importacao_valores_itens(import_id):
    acesso_negado = _require_valores_access()
    if acesso_negado:
        return acesso_negado

    # Buscar dados do banco temporário
    try:
        from sqlalchemy import text
        result = db.session.execute(
            text("SELECT dados FROM import_valores_temp WHERE id = :id"),
            {"id": import_id}
        )
        row = result.fetchone()
        if not row:
            flash('Importação não encontrada ou expirou.', 'warning')
            return redirect(url_for('itens.importar_planilha_valores_itens'))
        
        import json
        ok_rows = json.loads(row[0])

    except Exception as e:
        flash(f'Erro ao carregar dados da importação: {e}', 'danger')
        return redirect(url_for('itens.importar_planilha_valores_itens'))

    try:
        atualizados = 0
        itens_com_erro = []
        for row in ok_rows:
            item = db.session.get(Item, row['item_id'])
            if not item:
                itens_com_erro.append(f'Item não encontrado: {row.get("nome", "desconhecido")}')
                continue
            try:
                houve_alteracao = False
                if row.get('valor_item_informado'):
                    valor_novo = float(row['valor_item'] or 0)
                    if abs(float(item.valor_item or 0) - valor_novo) > 0.001:
                        item.valor_item = valor_novo
                        houve_alteracao = True
                
                if row.get('valor_material_informado'):
                    valor_novo = float(row.get('valor_material') or 0)
                    if abs(float(item.valor_material or 0) - valor_novo) > 0.001:
                        item.valor_material = valor_novo
                        houve_alteracao = True
                
                if row.get('outros_custos_informado'):
                    valor_novo = float(row.get('outros_custos') or 0)
                    if abs(float(item.outros_custos or 0) - valor_novo) > 0.001:
                        item.outros_custos = valor_novo
                        houve_alteracao = True
                
                if row.get('imposto_percentual_informado'):
                    valor_novo = float(row.get('imposto_percentual') or 0)
                    if abs(float(item.imposto_percentual or 0) - valor_novo) > 0.001:
                        item.imposto_percentual = valor_novo
                        houve_alteracao = True
                
                if houve_alteracao:
                    atualizados += 1
                
            except Exception as e:
                itens_com_erro.append(f'{row.get("nome", "desconhecido")}: {e}')
                continue

        db.session.commit()

        db.session.execute(
            text("DELETE FROM import_valores_temp WHERE id = :id"),
            {"id": import_id}
        )
        db.session.commit()

        flash(f'🎉 Importação concluída! {atualizados} itens atualizados com sucesso!', 'success')
        if atualizados == 0:
            flash('ℹ️ Nenhum valor foi alterado. Todos os valores da planilha já eram iguais aos do sistema.', 'info')
        if itens_com_erro:
            total_erros = len(itens_com_erro)
            preview_erros = '; '.join(itens_com_erro[:5])
            if total_erros > 5:
                preview_erros += f'; e mais {total_erros - 5} erro(s)'
            flash(f'⚠️ Alguns itens não puderam ser atualizados: {preview_erros}', 'warning')
        
        # Limpar sessão para evitar problema de cookie grande
        session.pop('import_valores_itens_rows', None)
        
        return redirect(url_for('itens.listar_valores_itens'))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao importar valores: {e}', 'danger')
        return redirect(url_for('itens.importar_planilha_valores_itens'))

@itens.route('/itens/novo', methods=['GET', 'POST'])
def novo_item():
    """Rota para cadastrar um novo item"""
    if request.method == 'POST':
        pode_ver_valores = _usuario_pode_ver_valores()
        # Validação de dados
        errors = validate_form_data(request.form, ['nome'])
        if errors:
            for error in errors:
                flash(error, 'danger')
            materiais = Material.query.all()
            trabalhos = Trabalho.query.all()
            return render_template('itens/novo.html', materiais=materiais, trabalhos=trabalhos, item=None, pode_ver_valores=pode_ver_valores)
        
        nome = request.form['nome']
        
        # Verificar se já existe um item com o mesmo nome
        item_existente = Item.query.filter_by(nome=nome).first()
        if item_existente:
            flash('Já existe um item com este nome!', 'danger')
            materiais = Material.query.all()
            trabalhos = Trabalho.query.all()
            return render_template('itens/novo.html', materiais=materiais, trabalhos=trabalhos, item=None, pode_ver_valores=pode_ver_valores)
        
        # Gerar código ACB automaticamente
        novo_codigo = generate_next_code(Item, "ACB", "codigo_acb")
        
        tipo_item = (request.form.get('tipo_item', 'producao') or 'producao').strip().lower()
        categoria_montagem = (request.form.get('categoria_montagem') or '').strip() or None

        # Criar o item
        item = Item(
            nome=nome,
            codigo_acb=novo_codigo,
            tipo_item=tipo_item,
            categoria_montagem=categoria_montagem,
            tipo_bruto=request.form.get('tipo_bruto', ''),
            tamanho_peca=request.form.get('tamanho_peca', ''),
            tempera='tempera' in request.form,
            tipo_tempera=request.form.get('tipo_tempera', ''),
            retifica='retifica' in request.form,
            pintura='pintura' in request.form,
            tipo_pintura=request.form.get('tipo_pintura', ''),
            cor_pintura=request.form.get('cor_pintura', ''),
            oleo_protetivo='oleo_protetivo' in request.form,
            zincagem='zincagem' in request.form,
            tipo_zincagem=request.form.get('tipo_zincagem', ''),
            tipo_embalagem=request.form.get('tipo_embalagem', '')
        )

        if pode_ver_valores:
            try:
                _apply_campos_financeiros(item, request.form)
            except ValueError as e:
                flash(str(e), 'danger')
                materiais = Material.query.all()
                trabalhos = Trabalho.query.all()
                return render_template('itens/novo.html', materiais=materiais, trabalhos=trabalhos, item=None, pode_ver_valores=pode_ver_valores)

        if item.tipo_item == 'montagem':
            item.eh_composto = False
        
        # Validar e converter o peso
        try:
            peso = request.form.get('peso', 0)
            item.peso = float(peso) if peso else 0
        except ValueError:
            flash('O peso deve ser um número válido', 'danger')
            materiais = Material.query.all()
            trabalhos = Trabalho.query.all()
            return render_template('itens/novo.html', materiais=materiais, trabalhos=trabalhos, item=None, pode_ver_valores=pode_ver_valores)
        
        # Upload de arquivos
        if 'desenho_tecnico' in request.files:
            item.desenho_tecnico = save_file(request.files['desenho_tecnico'], 'desenhos')
        
        if 'imagem' in request.files:
            item.imagem = save_file(request.files['imagem'], 'imagens')
        
        if 'instrucoes_trabalho' in request.files:
            item.instrucoes_trabalho = save_file(request.files['instrucoes_trabalho'], 'instrucoes')
        
        db.session.add(item)
        db.session.commit()
        
        if item.tipo_item != 'montagem':
            # Adicionar materiais
            materiais_json = parse_json_field(request.form, 'materiais')
            for mat in materiais_json:
                try:
                    item_material = ItemMaterial(
                        item_id=item.id,
                        material_id=mat['id'],
                        comprimento=float(mat.get('comprimento', 0) or 0),
                        quantidade=int(mat.get('quantidade', 1) or 1)
                    )
                    db.session.add(item_material)
                except (ValueError, KeyError) as e:
                    flash(f'Erro ao processar material: {str(e)}', 'danger')

            # Adicionar trabalhos
            trabalhos_json = parse_json_field(request.form, 'trabalhos')
            for trab in trabalhos_json:
                try:
                    item_trabalho = ItemTrabalho(
                        item_id=item.id,
                        trabalho_id=trab['id'],
                        tempo_setup=int(trab.get('tempo_setup', 0) or 0),
                        tempo_peca=int(trab.get('tempo_peca', 0) or 0)
                    )
                    db.session.add(item_trabalho)
                except (ValueError, KeyError) as e:
                    flash(f'Erro ao processar trabalho: {str(e)}', 'danger')
        
        # Adicionar arquivos CNC
        if 'cnc_files' in request.files:
            cnc_files = request.files.getlist('cnc_files')
            maquina = request.form.get('maquina_cnc', '')
            criador_id = session.get('usuario_id')
            if not criador_id:
                flash('Por favor, faça login para adicionar arquivos CNC.', 'warning')
                return redirect(url_for('auth.login', next=request.url))
            for file in cnc_files:
                if file and file.filename.endswith(('.txt', '.nc')):
                    filename = save_file(file, 'cnc_files')
                    arquivo_cnc = ArquivoCNC(
                        item_id=item.id,
                        nome_arquivo=filename,
                        caminho_arquivo=os.path.join('uploads/cnc_files', filename),
                        maquina=maquina,
                        criador_id=criador_id
                    )
                    db.session.add(arquivo_cnc)
        
        db.session.commit()
        flash('Item cadastrado com sucesso!', 'success')
        return redirect(url_for('itens.listar_itens'))
    
    materiais = Material.query.all()
    trabalhos = Trabalho.query.all()
    tipo_item_default = (request.args.get('tipo_item') or 'producao').strip().lower()
    if tipo_item_default not in ('producao', 'montagem'):
        tipo_item_default = 'producao'
    return render_template('itens/novo.html', materiais=materiais, trabalhos=trabalhos, item=None, tipo_item_default=tipo_item_default, pode_ver_valores=_usuario_pode_ver_valores())

@itens.route('/itens/editar/<int:item_id>', methods=['GET', 'POST'])
def editar_item(item_id):
    """Rota para editar um item existente"""
    item = Item.query.get_or_404(item_id)
    materiais = Material.query.all()
    trabalhos = Trabalho.query.all()
    pode_ver_valores = _usuario_pode_ver_valores()

    item_materiais = []
    for im in item.materiais:
        material = Material.query.get(im.material_id)
        item_materiais.append({
            'id': im.material_id,
            'nome': material.nome,
            'tipo': material.tipo,
            'comprimento': im.comprimento,
            'quantidade': im.quantidade
        })

    item_trabalhos = []
    for it in item.trabalhos:
        trabalho = Trabalho.query.get(it.trabalho_id)
        item_trabalhos.append({
            'id': it.trabalho_id,
            'nome': trabalho.nome,
            'tempo_setup': it.tempo_setup,
            'tempo_peca': it.tempo_peca
        })
    
    if request.method == 'POST':
        # Validação de dados
        errors = validate_form_data(request.form, ['nome'])
        if errors:
            for error in errors:
                flash(error, 'danger')
            return render_template('itens/editar.html', item=item, materiais=materiais, trabalhos=trabalhos, pode_ver_valores=pode_ver_valores, item_materiais=item_materiais, item_trabalhos=item_trabalhos)
        
        nome = request.form['nome']
        
        # Verificar se já existe outro item com o mesmo nome (exceto o atual)
        item_existente = Item.query.filter(Item.nome == nome, Item.id != item_id).first()
        if item_existente:
            flash('Já existe um item com este nome!', 'danger')
            return render_template('itens/editar.html', item=item, materiais=materiais, trabalhos=trabalhos, pode_ver_valores=pode_ver_valores, item_materiais=item_materiais, item_trabalhos=item_trabalhos)
        
        tipo_item = (request.form.get('tipo_item', item.tipo_item or 'producao') or 'producao').strip().lower()
        categoria_montagem = (request.form.get('categoria_montagem') or '').strip() or None

        # Atualizar dados do item
        item.nome = nome
        item.tipo_item = tipo_item
        item.categoria_montagem = categoria_montagem
        item.tipo_bruto = request.form.get('tipo_bruto', '')
        item.tamanho_peca = request.form.get('tamanho_peca', '')
        item.tempera = 'tempera' in request.form
        item.tipo_tempera = request.form.get('tipo_tempera', '')
        item.retifica = 'retifica' in request.form
        item.pintura = 'pintura' in request.form
        item.tipo_pintura = request.form.get('tipo_pintura', '')
        item.cor_pintura = request.form.get('cor_pintura', '')
        item.oleo_protetivo = 'oleo_protetivo' in request.form
        item.zincagem = 'zincagem' in request.form
        item.tipo_zincagem = request.form.get('tipo_zincagem', '')
        item.tipo_embalagem = request.form.get('tipo_embalagem', '')

        if pode_ver_valores:
            try:
                _apply_campos_financeiros(item, request.form)
            except ValueError as e:
                flash(str(e), 'danger')
                return render_template('itens/editar.html', item=item, materiais=materiais, trabalhos=trabalhos, pode_ver_valores=pode_ver_valores, item_materiais=item_materiais, item_trabalhos=item_trabalhos)
        
        # Validar e converter o peso
        try:
            peso = request.form.get('peso', 0)
            item.peso = float(peso) if peso else 0
        except ValueError:
            flash('O peso deve ser um número válido', 'danger')
            return render_template('itens/editar.html', item=item, materiais=materiais, trabalhos=trabalhos, pode_ver_valores=pode_ver_valores, item_materiais=item_materiais, item_trabalhos=item_trabalhos)
        
        # Upload de arquivos
        if 'desenho_tecnico' in request.files and request.files['desenho_tecnico'].filename:
            item.desenho_tecnico = save_file(request.files['desenho_tecnico'], 'desenhos')
        
        if 'imagem' in request.files and request.files['imagem'].filename:
            item.imagem = save_file(request.files['imagem'], 'imagens')
        
        if 'instrucoes_trabalho' in request.files and request.files['instrucoes_trabalho'].filename:
            item.instrucoes_trabalho = save_file(request.files['instrucoes_trabalho'], 'instrucoes')

        if getattr(item, 'criado_via_importacao_estoque', False):
            item.criado_via_importacao_estoque = False

        if item.tipo_item == 'montagem':
            item.eh_composto = False
            ItemMaterial.query.filter_by(item_id=item.id).delete()
            ItemTrabalho.query.filter_by(item_id=item.id).delete()
        else:
            # Atualizar materiais
            materiais_json = parse_json_field(request.form, 'materiais')

            # Remover materiais existentes
            ItemMaterial.query.filter_by(item_id=item.id).delete()

            # Adicionar novos materiais
            for mat in materiais_json:
                try:
                    item_material = ItemMaterial(
                        item_id=item.id,
                        material_id=mat['id'],
                        comprimento=float(mat.get('comprimento', 0) or 0),
                        quantidade=int(mat.get('quantidade', 1) or 1)
                    )
                    db.session.add(item_material)
                except (ValueError, KeyError) as e:
                    flash(f'Erro ao processar material: {str(e)}', 'danger')

            # Atualizar trabalhos
            trabalhos_json = parse_json_field(request.form, 'trabalhos')

            # Remover trabalhos existentes
            ItemTrabalho.query.filter_by(item_id=item.id).delete()

            # Adicionar novos trabalhos
            for trab in trabalhos_json:
                try:
                    item_trabalho = ItemTrabalho(
                        item_id=item.id,
                        trabalho_id=trab['id'],
                        tempo_setup=int(trab.get('tempo_setup', 0) or 0),
                        tempo_peca=int(trab.get('tempo_peca', 0) or 0)
                    )
                    db.session.add(item_trabalho)
                except (ValueError, KeyError) as e:
                    flash(f'Erro ao processar trabalho: {str(e)}', 'danger')
        
        # Adicionar arquivos CNC
        if 'cnc_files' in request.files:
            cnc_files = request.files.getlist('cnc_files')
            maquina = request.form.get('maquina_cnc', '')
            criador_id = session.get('usuario_id')
            if not criador_id:
                flash('Por favor, faça login para adicionar arquivos CNC.', 'warning')
                return redirect(url_for('auth.login', next=request.url))
            for file in cnc_files:
                if file and file.filename.endswith(('.txt', '.nc')):
                    filename = save_file(file, 'cnc_files')
                    arquivo_cnc = ArquivoCNC(
                        item_id=item.id,
                        nome_arquivo=filename,
                        caminho_arquivo=os.path.join('uploads/cnc_files', filename),
                        maquina=maquina,
                        criador_id=criador_id
                    )
                    db.session.add(arquivo_cnc)
        
        db.session.commit()
        flash('Item atualizado com sucesso!', 'success')
        return redirect(url_for('itens.listar_itens'))
    
    return render_template('itens/editar.html', 
                          item=item, 
                          materiais=materiais, 
                          trabalhos=trabalhos,
                          pode_ver_valores=pode_ver_valores,
                          item_materiais=item_materiais,
                          item_trabalhos=item_trabalhos)

@itens.route('/itens/atualizar_ordem', methods=['POST'])
def atualizar_ordem():
    # Apenas retorna sucesso sem tentar salvar no banco
    # A ordem será mantida no localStorage do navegador
    return jsonify({'status': 'sucesso'})

@itens.route('/itens/<int:item_id>/adicionar-arquivo-cnc', methods=['POST'])
def adicionar_arquivo_cnc(item_id):
    """Rota para adicionar um arquivo CNC a um item existente"""
    item = Item.query.get_or_404(item_id)
    
    # Processar arquivos CNC enviados via formulário dinâmico
    cnc_files = request.files.getlist('cnc_files[]') if 'cnc_files[]' in request.files else []
    cnc_nomes = request.form.getlist('cnc_nomes[]') if 'cnc_nomes[]' in request.form else []
    cnc_maquinas = request.form.getlist('cnc_maquinas[]') if 'cnc_maquinas[]' in request.form else []
    
    # Criar diretório para arquivos CNC se não existir
    upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER_INSTRUCOES'], 'arquivos_cnc')
    os.makedirs(upload_dir, exist_ok=True)
    
    arquivos_adicionados = 0
    
    if cnc_files and len(cnc_files) > 0 and cnc_files[0].filename != '':
        # Processar cada arquivo CNC com seus metadados
        for i, file in enumerate(cnc_files):
            if file and file.filename.endswith(('.txt', '.nc')):
                # Obter nome e máquina correspondentes
                nome_personalizado = cnc_nomes[i] if i < len(cnc_nomes) else ''
                maquina = cnc_maquinas[i] if i < len(cnc_maquinas) else ''
                
                # Se não houver nome personalizado, usar o nome original
                if not nome_personalizado:
                    nome_personalizado = os.path.splitext(file.filename)[0]
                
                # Gerar nome de arquivo com extensão original
                extensao = os.path.splitext(file.filename)[1]
                nome_arquivo = secure_filename(f"{nome_personalizado}{extensao}")
                
                # Salvar arquivo
                file_path = os.path.join(upload_dir, nome_arquivo)
                file.save(file_path)
                
                # Criar registro no banco
                criador_id = session.get('usuario_id')
                if not criador_id:
                    flash('Por favor, faça login para adicionar arquivos CNC.', 'warning')
                    return redirect(url_for('auth.login', next=request.url))
                arquivo_cnc = ArquivoCNC(
                    item_id=item.id,
                    nome_arquivo=nome_arquivo,
                    caminho_arquivo=file_path,  # Adicionar o caminho do arquivo
                    maquina=maquina,
                    criador_id=criador_id
                )
                db.session.add(arquivo_cnc)
                arquivos_adicionados += 1
    
    if arquivos_adicionados > 0:
        db.session.commit()
        flash(f'{arquivos_adicionados} arquivo(s) CNC adicionado(s) com sucesso!', 'success')
    else:
        flash('Nenhum arquivo CNC válido foi enviado.', 'warning')
    
    return redirect(url_for('itens.editar_item', item_id=item_id))

@itens.route('/itens/<int:item_id>/remover-arquivo-cnc/<int:arquivo_id>', methods=['GET'])
def remover_arquivo_cnc(item_id, arquivo_id):
    """Rota para remover um arquivo CNC de um item"""
    arquivo = ArquivoCNC.query.get_or_404(arquivo_id)
    
    # Verificar se o arquivo pertence ao item correto
    if arquivo.item_id != item_id:
        flash('Arquivo não pertence a este item!', 'danger')
        return redirect(url_for('itens.editar_item', item_id=item_id))
    
    # Tentar excluir o arquivo físico
    try:
        if os.path.exists(arquivo.caminho_arquivo):
            os.remove(arquivo.caminho_arquivo)
    except Exception as e:
        flash(f'Erro ao excluir o arquivo: {str(e)}', 'warning')
    
    # Excluir o registro no banco de dados
    db.session.delete(arquivo)
    db.session.commit()
    
    flash('Arquivo CNC removido com sucesso!', 'success')
    return redirect(url_for('itens.editar_item', item_id=item_id))

@itens.route('/itens/download-arquivo-cnc/<int:arquivo_id>', methods=['GET'])
def download_arquivo_cnc(arquivo_id):
    """Rota para download de arquivo CNC"""
    arquivo = ArquivoCNC.query.get_or_404(arquivo_id)
    
    # Verificar se o arquivo existe
    if not os.path.exists(arquivo.caminho_arquivo):
        flash('Arquivo não encontrado no servidor.', 'danger')
        return redirect(url_for('itens.editar_item', item_id=arquivo.item_id))
    
    # Retornar o arquivo para download
    return send_file(arquivo.caminho_arquivo, as_attachment=True, download_name=arquivo.nome_arquivo)

@itens.route('/itens/excluir/<int:item_id>', methods=['POST'])
def excluir_item(item_id):
    """Rota para excluir um item"""
    item = Item.query.get_or_404(item_id)
    
    # Verificar se o item está associado a algum pedido
    pedidos_associados = Pedido.query.filter_by(item_id=item_id).count()
    if pedidos_associados > 0:
        flash('Não é possível excluir um item associado a pedidos!', 'danger')
        return redirect(url_for('itens.listar_itens'))
    
    # Excluir relações com materiais e trabalhos
    ItemMaterial.query.filter_by(item_id=item_id).delete()
    ItemTrabalho.query.filter_by(item_id=item_id).delete()
    ArquivoCNC.query.filter_by(item_id=item_id).delete()
    
    # Armazenar o nome para mensagem
    nome_item = item.nome
    
    # Excluir o item
    db.session.delete(item)
    db.session.commit()
    
    flash(f'Item "{nome_item}" excluído com sucesso!', 'success')
    return redirect(url_for('itens.listar_itens'))

@itens.route('/itens/visualizar/<int:item_id>')
def visualizar_item(item_id):
    """Rota para visualizar detalhes de um item"""
    item = Item.query.get_or_404(item_id)
    return render_template('itens/visualizar.html', item=item, pode_ver_valores=_usuario_pode_ver_valores())


def _build_supabase_public_url_from_file_path(file_path: str) -> str | None:
    if not file_path:
        return None

    if not (file_path.startswith('supabase://') or file_path.startswith('supabase:/')):
        return None

    import os
    from urllib.parse import quote

    supabase_url = os.environ.get('SUPABASE_URL', '').rstrip('/')
    bucket_env = os.environ.get('SUPABASE_BUCKET', 'uploads')
    if not supabase_url:
        return None

    file_name = file_path.replace('supabase://', '').replace('supabase:/', '').lstrip('/')
    KNOWN_FOLDERS = {'imagens', 'desenhos', 'instrucoes', 'cnc_files', 'maquinas', 'castanhas', 'gabaritos', 'folhas_processo'}
    parts = file_name.split('/', 1)
    if len(parts) > 1 and parts[0] not in KNOWN_FOLDERS:
        bucket = parts[0]
        rel_path = parts[1]
    else:
        bucket = bucket_env
        rel_path = file_name

    rel_encoded = quote(rel_path, safe='/')
    return f"{supabase_url}/storage/v1/object/public/{bucket}/{rel_encoded}"


def _get_item_desenho_pdf_bytes(item: Item) -> tuple[bytes | None, str | None]:
    file_path = item.desenho_tecnico
    if not file_path:
        return None, None

    if file_path.startswith('http://') or file_path.startswith('https://'):
        import requests
        resp = requests.get(file_path, timeout=30)
        if resp.status_code != 200:
            return None, None
        return resp.content, 'application/pdf'

    if file_path.startswith('supabase://') or file_path.startswith('supabase:/'):
        import requests
        public_url = _build_supabase_public_url_from_file_path(file_path)
        if not public_url:
            return None, None
        resp = requests.get(public_url, timeout=30)
        if resp.status_code != 200:
            return None, None
        return resp.content, 'application/pdf'

    normalized = file_path.replace('\\', '/').lstrip('/')
    if normalized.startswith('uploads/'):
        normalized = normalized[len('uploads/'):]

    if normalized.startswith('desenhos/'):
        filename = normalized.split('/', 1)[1]
        local_path = os.path.join(current_app.config['UPLOAD_FOLDER_DESENHOS'], filename)
        if not os.path.exists(local_path):
            return None, None
        with open(local_path, 'rb') as f:
            return f.read(), 'application/pdf'

    return None, None


def _stamp_pdf_approved(pdf_bytes: bytes, aprovado_por_nome: str | None, aprovado_em: datetime | None) -> bytes:
    from pypdf import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas
    from reportlab.lib.colors import HexColor

    reader = PdfReader(io.BytesIO(pdf_bytes))
    writer = PdfWriter()

    nome = (aprovado_por_nome or '').strip()
    data_str = aprovado_em.strftime('%d/%m/%Y %H:%M') if aprovado_em else ''

    for page in reader.pages:
        mb = page.mediabox
        width = float(mb.width)
        height = float(mb.height)

        overlay_stream = io.BytesIO()
        c = canvas.Canvas(overlay_stream, pagesize=(width, height))

        c.saveState()
        c.setStrokeColor(HexColor('#198754'))
        c.setFillColor(HexColor('#198754'))
        c.setLineWidth(3)

        stamp_w = 180
        stamp_h = 60
        margin = 40
        x = width - margin - stamp_w
        y = height - margin - stamp_h

        c.translate(x + stamp_w / 2, y + stamp_h / 2)
        c.rotate(6)
        c.roundRect(-stamp_w / 2, -stamp_h / 2, stamp_w, stamp_h, 8, stroke=1, fill=0)
        c.setFont('Helvetica-Bold', 16)
        c.drawCentredString(0, 10, 'APROVADO')
        c.setFont('Helvetica-Bold', 9)
        if nome:
            c.drawCentredString(0, -4, nome)
        if data_str:
            c.setFont('Helvetica', 8)
            c.drawCentredString(0, -18, data_str)
        c.restoreState()

        c.showPage()
        c.save()
        overlay_stream.seek(0)
        overlay_pdf = PdfReader(overlay_stream)
        overlay_page = overlay_pdf.pages[0]

        page.merge_page(overlay_page)
        writer.add_page(page)

    out = io.BytesIO()
    writer.write(out)
    out.seek(0)
    return out.read()


@itens.route('/itens/desenho-pdf/<int:item_id>')
def desenho_pdf_item(item_id):
    item = Item.query.get_or_404(item_id)
    if not item.desenho_tecnico:
        flash('Este item não possui desenho técnico.', 'warning')
        return redirect(url_for('itens.visualizar_item', item_id=item.id))

    if not item.desenho_aprovado_em:
        return redirect(item.desenho_tecnico_path)

    pdf_bytes, _mime = _get_item_desenho_pdf_bytes(item)
    if not pdf_bytes:
        flash('Não foi possível obter o PDF do desenho para carimbar.', 'danger')
        return redirect(item.desenho_tecnico_path)

    try:
        stamped = _stamp_pdf_approved(pdf_bytes, item.desenho_aprovado_por_nome, item.desenho_aprovado_em)
    except Exception:
        flash('Erro ao gerar PDF carimbado.', 'danger')
        return redirect(item.desenho_tecnico_path)

    filename = f"{(item.codigo_acb or 'DESENHO')}_{(item.nome or 'item')}.pdf".replace(' ', '_')
    return send_file(
        io.BytesIO(stamped),
        mimetype='application/pdf',
        as_attachment=False,
        download_name=filename,
    )


@itens.route('/itens/imprimir-desenho/<int:item_id>')
def imprimir_desenho_item(item_id):
    """Impressão do desenho técnico do Item com carimbo de aprovação."""
    item = Item.query.get_or_404(item_id)
    return redirect(url_for('itens.desenho_pdf_item', item_id=item.id))


@itens.route('/itens/aprovar-desenho/<int:item_id>', methods=['POST'])
def aprovar_desenho_item(item_id):
    item = Item.query.get_or_404(item_id)
    usuario = getattr(g, 'usuario', None)
    if not usuario or getattr(usuario, 'nivel_acesso', None) != 'admin':
        flash('Você não tem permissão para aprovar o desenho.', 'danger')
        return redirect(url_for('itens.visualizar_item', item_id=item.id))

    if not item.desenho_tecnico:
        flash('Este item não possui desenho técnico para aprovar.', 'warning')
        return redirect(url_for('itens.visualizar_item', item_id=item.id))

    item.desenho_aprovado_em = datetime.utcnow()
    item.desenho_aprovado_por_id = usuario.id
    item.desenho_aprovado_por_nome = getattr(usuario, 'nome', None)
    db.session.commit()
    flash('Desenho aprovado.', 'success')
    return redirect(url_for('itens.visualizar_item', item_id=item.id))


@itens.route('/itens/desaprovar-desenho/<int:item_id>', methods=['POST'])
def desaprovar_desenho_item(item_id):
    item = Item.query.get_or_404(item_id)
    usuario = getattr(g, 'usuario', None)
    if not usuario or getattr(usuario, 'nivel_acesso', None) != 'admin':
        flash('Você não tem permissão para remover aprovação do desenho.', 'danger')
        return redirect(url_for('itens.visualizar_item', item_id=item.id))

    item.desenho_aprovado_em = None
    item.desenho_aprovado_por_id = None
    item.desenho_aprovado_por_nome = None
    db.session.commit()
    flash('Aprovação do desenho removida.', 'success')
    return redirect(url_for('itens.visualizar_item', item_id=item.id))

@itens.route('/api/item/<int:item_id>')
def api_item(item_id):
    """API para obter dados de um item"""
    item = Item.query.get_or_404(item_id)
    return jsonify({
        'id': item.id,
        'nome': item.nome,
        'codigo_acb': item.codigo_acb
    })

# ==========================================
# ROTAS PARA ITENS COMPOSTOS
# ==========================================

@itens.route('/itens/composto/novo', methods=['GET', 'POST'])
def novo_item_composto():
    """Rota para cadastrar um novo item composto"""
    if request.method == 'POST':
        pode_ver_valores = _usuario_pode_ver_valores()
        # Validação de dados
        errors = validate_form_data(request.form, ['nome'])
        if errors:
            for error in errors:
                flash(error, 'danger')
            itens_disponiveis = Item.query.filter_by(eh_composto=False).all()
            return render_template('itens/composto_novo.html', itens_disponiveis=itens_disponiveis, pode_ver_valores=pode_ver_valores)
        
        nome = request.form['nome']
        
        # Verificar se já existe um item com o mesmo nome
        item_existente = Item.query.filter_by(nome=nome).first()
        if item_existente:
            flash('Já existe um item com este nome!', 'danger')
            itens_disponiveis = Item.query.filter_by(eh_composto=False).all()
            return render_template('itens/composto_novo.html', itens_disponiveis=itens_disponiveis, pode_ver_valores=pode_ver_valores)
        
        # Gerar código ACB automaticamente
        novo_codigo = generate_next_code(Item, "ACB", "codigo_acb")
        
        # Criar o item composto
        item = Item(
            nome=nome,
            codigo_acb=novo_codigo,
            eh_composto=True,  # Marcar como item composto
            tipo_bruto=request.form.get('tipo_bruto', ''),
            tamanho_peca=request.form.get('tamanho_peca', ''),
            tipo_embalagem=request.form.get('tipo_embalagem', ''),
            tempera='tempera' in request.form,
            tipo_tempera=request.form.get('tipo_tempera', ''),
            retifica='retifica' in request.form,
            pintura='pintura' in request.form,
            tipo_pintura=request.form.get('tipo_pintura', ''),
            cor_pintura=request.form.get('cor_pintura', ''),
            oleo_protetivo='oleo_protetivo' in request.form,
            zincagem='zincagem' in request.form,
            tipo_zincagem=request.form.get('tipo_zincagem', '')
        )

        if pode_ver_valores:
            try:
                _apply_campos_financeiros(item, request.form)
            except ValueError as e:
                flash(str(e), 'danger')
                itens_disponiveis = Item.query.filter_by(eh_composto=False).all()
                return render_template('itens/composto_novo.html', itens_disponiveis=itens_disponiveis, pode_ver_valores=pode_ver_valores)
        
        # Upload de arquivos
        if 'desenho_tecnico' in request.files:
            item.desenho_tecnico = save_file(request.files['desenho_tecnico'], 'desenhos')
        
        if 'imagem' in request.files:
            item.imagem = save_file(request.files['imagem'], 'imagens')
        
        if 'instrucoes_trabalho' in request.files:
            item.instrucoes_trabalho = save_file(request.files['instrucoes_trabalho'], 'instrucoes')
        
        db.session.add(item)
        db.session.commit()
        
        # Adicionar componentes
        componentes_json = parse_json_field(request.form, 'componentes')
        componentes_map = {}
        for comp in componentes_json:
            try:
                # Verificar se não está tentando adicionar o próprio item como componente
                if int(comp['id']) == item.id:
                    flash('Um item não pode ser componente de si mesmo!', 'danger')
                    continue

                componente_id = int(comp['id'])
                componente = Item.query.get(componente_id)
                if not componente:
                    continue

                comprimento_mm = comp.get('comprimento_mm', None)
                if comprimento_mm is not None and str(comprimento_mm).strip() == '':
                    comprimento_mm = None
                if comprimento_mm is not None:
                    try:
                        comprimento_mm = float(str(comprimento_mm).replace(',', '.'))
                    except ValueError:
                        comprimento_mm = None

                is_chaveta = (componente.categoria_montagem or '').strip().lower() == 'chaveta'
                if not is_chaveta:
                    comprimento_mm = None

                qtd = int(comp.get('quantidade', 1) or 1)

                if is_chaveta:
                    entry = componentes_map.get(componente_id)
                    if not entry:
                        componentes_map[componente_id] = {
                            'quantidade': 1,
                            'comprimento_mm': (comprimento_mm or 0.0) * qtd,
                            'observacoes': comp.get('observacoes', '')
                        }
                    else:
                        entry['comprimento_mm'] = (entry.get('comprimento_mm') or 0.0) + ((comprimento_mm or 0.0) * qtd)
                        obs = (comp.get('observacoes', '') or '').strip()
                        if obs:
                            prev = (entry.get('observacoes', '') or '').strip()
                            entry['observacoes'] = (prev + ' | ' + obs) if prev else obs
                else:
                    if componente_id in componentes_map:
                        continue
                    componentes_map[componente_id] = {
                        'quantidade': qtd,
                        'comprimento_mm': None,
                        'observacoes': comp.get('observacoes', '')
                    }
            except (ValueError, KeyError) as e:
                flash(f'Erro ao processar componente: {str(e)}', 'danger')

        for componente_id, payload in componentes_map.items():
            item_composto = ItemComposto(
                item_pai_id=item.id,
                item_componente_id=componente_id,
                quantidade=int(payload.get('quantidade', 1) or 1),
                comprimento_mm=payload.get('comprimento_mm', None),
                observacoes=payload.get('observacoes', '')
            )
            db.session.add(item_composto)
        
        db.session.commit()
        flash('Item composto cadastrado com sucesso!', 'success')
        return redirect(url_for('itens.listar_itens'))
    
    # GET - mostrar formulário
    itens_disponiveis = Item.query.filter_by(eh_composto=False).all()
    return render_template('itens/composto_novo.html', itens_disponiveis=itens_disponiveis, pode_ver_valores=_usuario_pode_ver_valores())

@itens.route('/itens/composto/editar/<int:item_id>', methods=['GET', 'POST'])
def editar_item_composto(item_id):
    """Rota para editar um item composto existente"""
    item = Item.query.get_or_404(item_id)
    pode_ver_valores = _usuario_pode_ver_valores()
    
    if not item.eh_composto:
        flash('Este item não é um item composto!', 'danger')
        return redirect(url_for('itens.listar_itens'))
    
    if request.method == 'POST':
        # Validação de dados
        errors = validate_form_data(request.form, ['nome'])
        if errors:
            for error in errors:
                flash(error, 'danger')
            itens_disponiveis = Item.query.filter_by(eh_composto=False).all()
            return render_template('itens/composto_editar.html', item=item, itens_disponiveis=itens_disponiveis, pode_ver_valores=pode_ver_valores)
        
        nome = request.form['nome']
        
        # Verificar se já existe outro item com o mesmo nome (exceto o atual)
        item_existente = Item.query.filter(Item.nome == nome, Item.id != item_id).first()
        if item_existente:
            flash('Já existe um item com este nome!', 'danger')
            itens_disponiveis = Item.query.filter_by(eh_composto=False).all()
            return render_template('itens/composto_editar.html', item=item, itens_disponiveis=itens_disponiveis, pode_ver_valores=pode_ver_valores)
        
        # Atualizar dados do item
        item.nome = nome
        item.tipo_bruto = request.form.get('tipo_bruto', '')
        item.tamanho_peca = request.form.get('tamanho_peca', '')
        item.tipo_embalagem = request.form.get('tipo_embalagem', '')
        item.tempera = 'tempera' in request.form
        item.tipo_tempera = request.form.get('tipo_tempera', '')
        item.retifica = 'retifica' in request.form
        item.pintura = 'pintura' in request.form
        item.tipo_pintura = request.form.get('tipo_pintura', '')
        item.cor_pintura = request.form.get('cor_pintura', '')
        item.oleo_protetivo = 'oleo_protetivo' in request.form
        item.zincagem = 'zincagem' in request.form
        item.tipo_zincagem = request.form.get('tipo_zincagem', '')

        if pode_ver_valores:
            try:
                _apply_campos_financeiros(item, request.form)
            except ValueError as e:
                flash(str(e), 'danger')
                itens_disponiveis = Item.query.filter_by(eh_composto=False).all()
                return render_template('itens/composto_editar.html', item=item, itens_disponiveis=itens_disponiveis, pode_ver_valores=pode_ver_valores)
        
        # Upload de arquivos
        if 'desenho_tecnico' in request.files and request.files['desenho_tecnico'].filename:
            item.desenho_tecnico = save_file(request.files['desenho_tecnico'], 'desenhos')
        
        if 'imagem' in request.files and request.files['imagem'].filename:
            item.imagem = save_file(request.files['imagem'], 'imagens')
        
        if 'instrucoes_trabalho' in request.files and request.files['instrucoes_trabalho'].filename:
            item.instrucoes_trabalho = save_file(request.files['instrucoes_trabalho'], 'instrucoes')
        
        # Remover componentes existentes
        ItemComposto.query.filter_by(item_pai_id=item.id).delete()
        
        # Adicionar novos componentes
        componentes_json = parse_json_field(request.form, 'componentes')
        componentes_map = {}
        for comp in componentes_json:
            try:
                # Verificar se não está tentando adicionar o próprio item como componente
                if int(comp['id']) == item.id:
                    flash('Um item não pode ser componente de si mesmo!', 'danger')
                    continue

                componente_id = int(comp['id'])
                componente = Item.query.get(componente_id)
                if not componente:
                    continue

                comprimento_mm = comp.get('comprimento_mm', None)
                if comprimento_mm is not None and str(comprimento_mm).strip() == '':
                    comprimento_mm = None
                if comprimento_mm is not None:
                    try:
                        comprimento_mm = float(str(comprimento_mm).replace(',', '.'))
                    except ValueError:
                        comprimento_mm = None

                is_chaveta = (componente.categoria_montagem or '').strip().lower() == 'chaveta'
                if not is_chaveta:
                    comprimento_mm = None

                qtd = int(comp.get('quantidade', 1) or 1)

                if is_chaveta:
                    entry = componentes_map.get(componente_id)
                    if not entry:
                        componentes_map[componente_id] = {
                            'quantidade': 1,
                            'comprimento_mm': (comprimento_mm or 0.0) * qtd,
                            'observacoes': comp.get('observacoes', '')
                        }
                    else:
                        entry['comprimento_mm'] = (entry.get('comprimento_mm') or 0.0) + ((comprimento_mm or 0.0) * qtd)
                        obs = (comp.get('observacoes', '') or '').strip()
                        if obs:
                            prev = (entry.get('observacoes', '') or '').strip()
                            entry['observacoes'] = (prev + ' | ' + obs) if prev else obs
                else:
                    if componente_id in componentes_map:
                        continue
                    componentes_map[componente_id] = {
                        'quantidade': qtd,
                        'comprimento_mm': None,
                        'observacoes': comp.get('observacoes', '')
                    }
            except (ValueError, KeyError) as e:
                flash(f'Erro ao processar componente: {str(e)}', 'danger')

        for componente_id, payload in componentes_map.items():
            item_composto = ItemComposto(
                item_pai_id=item.id,
                item_componente_id=componente_id,
                quantidade=int(payload.get('quantidade', 1) or 1),
                comprimento_mm=payload.get('comprimento_mm', None),
                observacoes=payload.get('observacoes', '')
            )
            db.session.add(item_composto)
        
        db.session.commit()
        flash('Item composto atualizado com sucesso!', 'success')
        return redirect(url_for('itens.listar_itens'))
    
    # GET - mostrar formulário preenchido
    itens_disponiveis = Item.query.filter_by(eh_composto=False).all()
    
    # Obter componentes do item para preencher o formulário
    item_componentes = []
    for ic in item.componentes:
        componente = Item.query.get(ic.item_componente_id)
        item_componentes.append({
            'id': ic.item_componente_id,
            'nome': componente.nome,
            'codigo': componente.codigo_acb,
            'categoria_montagem': componente.categoria_montagem or '',
            'quantidade': ic.quantidade,
            'comprimento_mm': ic.comprimento_mm,
            'observacoes': ic.observacoes or ''
        })
    
    return render_template('itens/composto_editar.html', 
                          item=item, 
                          itens_disponiveis=itens_disponiveis,
                          item_componentes=item_componentes,
                          pode_ver_valores=pode_ver_valores)

@itens.route('/itens/composto/visualizar/<int:item_id>')
def visualizar_item_composto(item_id):
    """Rota para visualizar detalhes de um item composto"""
    item = Item.query.get_or_404(item_id)
    
    if not item.eh_composto:
        flash('Este item não é um item composto!', 'danger')
        return redirect(url_for('itens.listar_itens'))
    
    return render_template('itens/composto_visualizar.html', item=item, pode_ver_valores=_usuario_pode_ver_valores())


@itens.route('/itens/composto/imprimir/<int:item_id>')
def imprimir_item_composto(item_id):
    """Rota para imprimir a composição (BOM) de um item composto"""
    item = Item.query.get_or_404(item_id)

    if not item.eh_composto:
        flash('Este item não é um item composto!', 'danger')
        return redirect(url_for('itens.listar_itens'))

    return render_template('itens/composto_imprimir.html', item=item, now=datetime.now())

@itens.route('/api/itens/nao-compostos')
def api_itens_nao_compostos():
    """API para obter lista de itens que não são compostos (para usar como componentes)"""
    itens = Item.query.filter_by(eh_composto=False).all()
    return jsonify([{
        'id': item.id,
        'nome': item.nome,
        'codigo_acb': item.codigo_acb,
        'peso': item.peso or 0
    } for item in itens])
