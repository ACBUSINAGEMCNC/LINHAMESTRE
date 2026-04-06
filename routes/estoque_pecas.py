import json
import re
import unicodedata

from datetime import date, datetime

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app, session
from sqlalchemy import or_
from sqlalchemy.exc import ProgrammingError
from openpyxl import load_workbook

from models import db, EstoquePecas, Item, MovimentacaoEstoquePecas, EstoquePecasSlotTemp, Usuario
from utils import validate_form_data, generate_next_code, get_file_url

estoque_pecas = Blueprint('estoque_pecas', __name__)

# Constantes para verificação de admin master
ADMIN_MASTER_EMAIL = 'admin@acbusinagem.com.br'


def _require_acesso_valores():
    """Verifica se o usuário tem acesso a valores de itens"""
    if 'usuario_id' not in session:
        flash('Por favor, faça login para acessar esta página.', 'warning')
        return redirect(url_for('auth.login'))
    
    usuario = Usuario.query.get(session['usuario_id'])
    if not usuario:
        flash('Usuário não encontrado.', 'danger')
        return redirect(url_for('auth.login'))
    
    # Admin master sempre tem acesso
    if (getattr(usuario, 'email', '') or '').strip().lower() == ADMIN_MASTER_EMAIL and getattr(usuario, 'nivel_acesso', None) == 'admin':
        return None
    
    # Verificar acesso explícito aos valores
    if not bool(getattr(usuario, 'acesso_valores_itens', False)):
        flash('Você não tem permissão para acessar valores de itens.', 'danger')
        return redirect(url_for('index.index'))
    
    return None


def _parse_date_cell(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(v, fmt).date()
            except Exception:
                pass
    return None


def _normalize_header(value):
    if value is None:
        return ''
    s = str(value).strip().lower()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r'[^a-z0-9]+', '_', s)
    s = re.sub(r'_+', '_', s).strip('_')
    return s


def _canonical_estoque_excel_column(header: str) -> str:
    h = _normalize_header(header)
    aliases = {
        'nome': {'nome', 'item', 'nome_item', 'descricao', 'descricao_item', 'peca', 'nome_peca'},
        'quantidade': {'quantidade', 'qtd', 'qtde', 'saldo', 'estoque'},
        'observacao': {'observacao', 'obs', 'observacoes'},
        'data_entrada': {'data_entrada', 'entrada', 'data'},
    }
    for canonical, opts in aliases.items():
        if h in opts:
            return canonical
    return h


def _get_pending_import_rows():
    return session.get('import_estoque_excel_rows') or []


def _build_import_print_context(rows):
    existentes = [r for r in rows if not r.get('criar_item')]
    novos = [r for r in rows if r.get('criar_item')]
    return {
        'rows': rows,
        'existentes_count': len(existentes),
        'novos_count': len(novos),
        'total_count': len(rows),
        'total_quantidade': sum(int(r.get('quantidade') or 0) for r in rows),
        'now': datetime.now(),
    }


def _next_import_preview_code(counter):
    base = generate_next_code(Item, 'ACB', 'codigo_acb')
    try:
        prefix, number = base.split('-', 1)
        return f"{prefix}-{int(number) + counter:05d}"
    except Exception:
        return base


def _normalize_slots(slots, estante_padrao=None):
    out = []
    seen = set()
    for s in slots or []:
        try:
            est = int(s.get('estante') if isinstance(s, dict) else s[0])
            sec = int(s.get('secao') if isinstance(s, dict) else s[1])
            lin = int(s.get('linha') if isinstance(s, dict) else s[2])
            col = int(s.get('coluna') if isinstance(s, dict) else s[3])
        except Exception:
            continue
        if estante_padrao is not None:
            est = int(estante_padrao)
        if est < 1 or est > 8 or sec < 1 or sec > 4 or lin < 1 or lin > 2 or col < 1 or col > 6:
            continue
        key = (est, sec, lin, col)
        if key in seen:
            continue
        seen.add(key)
        out.append({'estante': est, 'secao': sec, 'linha': lin, 'coluna': col})
    out.sort(key=lambda x: (x['estante'], x['secao'], x['linha'], x['coluna']))
    return out


def _load_slots_json(raw, estante_padrao=None):
    if not raw:
        return []
    try:
        data = json.loads(raw)
    except Exception:
        return []
    return _normalize_slots(data, estante_padrao=estante_padrao)


def _dump_slots_json(slots):
    return json.dumps(_normalize_slots(slots), ensure_ascii=False)


def _slots_from_entity(entity):
    slots = _load_slots_json(getattr(entity, 'slots_json', None))
    if slots:
        return slots
    if getattr(entity, 'estante', None) and getattr(entity, 'secao', None) and getattr(entity, 'linha', None) and getattr(entity, 'coluna', None):
        return _normalize_slots([{
            'estante': int(entity.estante),
            'secao': int(entity.secao),
            'linha': int(entity.linha),
            'coluna': int(entity.coluna),
        }])
    return []

@estoque_pecas.route('/estoque-pecas')
def index():
    """Rota para a página principal do estoque de peças"""
    # Organizar estoque por prateleira
    show_zero = (request.args.get('show_zero') or '').strip().lower() in ('1', 'true', 'yes', 'sim')
    q = EstoquePecas.query
    if not show_zero:
        q = q.filter(EstoquePecas.quantidade > 0)
    estoque = q.order_by(EstoquePecas.estante, EstoquePecas.secao, EstoquePecas.linha, EstoquePecas.coluna).all()
    return render_template('estoque_pecas/index.html', estoque=estoque, show_zero=show_zero)


@estoque_pecas.route('/estoque-pecas/importar-excel', methods=['GET', 'POST'])
def importar_estoque_excel():
    if request.method == 'GET':
        session.pop('import_estoque_excel_rows', None)
        return render_template('estoque_pecas/importar_excel.html')

    arquivo = request.files.get('arquivo')
    if not arquivo or not arquivo.filename:
        flash('Selecione um arquivo XLSX para importar.', 'danger')
        return redirect(url_for('estoque_pecas.importar_estoque_excel'))

    try:
        wb = load_workbook(arquivo, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f'Erro ao ler XLSX: {e}', 'danger')
        return redirect(url_for('estoque_pecas.importar_estoque_excel'))

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header_row = row
        break
    if not header_row:
        flash('XLSX vazio ou sem cabeçalho.', 'danger')
        return redirect(url_for('estoque_pecas.importar_estoque_excel'))

    headers = [_canonical_estoque_excel_column(h) for h in header_row]
    col_index = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        col_index.setdefault(h, i)

    if 'nome' not in col_index and 'item' in col_index:
        col_index['nome'] = col_index['item']

    required = ['nome', 'quantidade']
    missing = [c for c in required if c not in col_index]
    if missing:
        encontrados = [h for h in headers if h]
        flash('Colunas obrigatórias ausentes no Excel: ' + ', '.join(missing), 'danger')
        flash('Cabeçalhos encontrados: ' + ', '.join(encontrados[:30]) + ('...' if len(encontrados) > 30 else ''), 'warning')
        return redirect(url_for('estoque_pecas.importar_estoque_excel'))

    preview_rows = []
    ok_rows = []
    ok_count = 0
    err_count = 0
    warn_count = 0
    skipped_empty_name_count = 0
    used_names = set()
    novos_counter = 0

    obs_col = col_index.get('observacao')
    data_col = col_index.get('data_entrada')

    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        nome_val = values[col_index['nome']] if col_index.get('nome') is not None else None
        quantidade_val = values[col_index['quantidade']] if col_index.get('quantidade') is not None else None
        observacao_val = values[obs_col] if obs_col is not None else None
        data_val = values[data_col] if data_col is not None else None

        errors = []
        warnings = []
        nome_s = str(nome_val or '').strip()
        nome_key = nome_s.casefold()
        observacao_s = str(observacao_val or '').strip()
        data_entrada = _parse_date_cell(data_val)

        if not any([nome_s, str(quantidade_val or '').strip(), observacao_s, data_val]):
            continue

        if not nome_s:
            skipped_empty_name_count += 1
            continue

        try:
            if isinstance(quantidade_val, float) and quantidade_val.is_integer():
                quantidade_val = int(quantidade_val)
            quantidade_int = int(str(quantidade_val).strip())
            if quantidade_int <= 0:
                errors.append('Quantidade deve ser > 0')
        except Exception:
            quantidade_int = None
            errors.append('Quantidade inválida')

        if nome_key and nome_key in used_names:
            warnings.append('Nome repetido na planilha: as quantidades serão importadas em linhas separadas')
        elif nome_key:
            used_names.add(nome_key)

        item = Item.query.filter(Item.nome == nome_s).first() if nome_s else None
        codigo_previsto = item.codigo_acb if item else _next_import_preview_code(novos_counter)
        if not item and len(errors) == 0:
            novos_counter += 1

        ok = len(errors) == 0
        if ok:
            ok_count += 1
            if warnings:
                warn_count += 1
            ok_rows.append({
                'row_number': row_number,
                'nome': nome_s,
                'quantidade': quantidade_int,
                'observacao': observacao_s,
                'data_entrada': data_entrada.isoformat() if data_entrada else None,
                'item_id': item.id if item else None,
                'codigo_acb': item.codigo_acb if item else None,
                'codigo_previsto': codigo_previsto,
                'criar_item': False if item else True,
                'acao_sugerida': 'usar_existente' if item else 'criar_novo',
                'selecionado': True,
            })
        else:
            err_count += 1

        preview_rows.append({
            'row_number': row_number,
            'nome': nome_s,
            'quantidade': '' if quantidade_int is None else quantidade_int,
            'observacao': observacao_s,
            'data_entrada': data_entrada.strftime('%d/%m/%Y') if data_entrada else '',
            'ok': ok,
            'errors': errors,
            'warnings': warnings,
            'encontrado': bool(item),
            'codigo_acb': item.codigo_acb if item else '',
            'codigo_previsto': codigo_previsto,
            'status_importacao': 'Cadastrar novo' if not item else 'Já cadastrado',
            'acao_sugerida': 'usar_existente' if item else 'criar_novo',
            'selecionado': True,
        })

    session['import_estoque_excel_rows'] = ok_rows
    if skipped_empty_name_count:
        flash(f'{skipped_empty_name_count} linha(s) sem nome foram ignoradas automaticamente na importação.', 'warning')
    return render_template(
        'estoque_pecas/importar_excel.html',
        preview_rows=preview_rows,
        preview_ok_count=ok_count,
        preview_error_count=err_count,
        preview_warning_count=warn_count,
    )


@estoque_pecas.route('/estoque-pecas/importar-excel/confirmar', methods=['POST'])
def confirmar_importacao_estoque_excel():
    ok_rows = _get_pending_import_rows()
    if not ok_rows:
        flash('Nenhuma importação pendente. Faça o upload do Excel novamente.', 'warning')
        return redirect(url_for('estoque_pecas.importar_estoque_excel'))

    try:
        importados = 0
        selecionados = []
        for idx, r in enumerate(ok_rows):
            marcado = request.form.get(f'selecionar_{idx}')
            if not marcado:
                continue
            acao = (request.form.get(f'acao_{idx}') or request.form.get('acao_global') or r.get('acao_sugerida') or '').strip()
            if acao == 'pular':
                continue

            r2 = dict(r)
            r2['acao_confirmada'] = acao
            selecionados.append(r2)

        if not selecionados:
            flash('Nenhuma linha foi selecionada para importação.', 'warning')
            return redirect(url_for('estoque_pecas.importar_estoque_excel'))

        for r in selecionados:
            item = Item.query.get(r.get('item_id')) if r.get('item_id') else None
            if r.get('acao_confirmada') == 'usar_existente' and not item:
                flash(f"A linha {r.get('row_number')} não possui item cadastrado para usar.", 'danger')
                return redirect(url_for('estoque_pecas.importar_estoque_excel'))

            if not item:
                item = Item(
                    nome=r['nome'],
                    codigo_acb=generate_next_code(Item, 'ACB', 'codigo_acb'),
                    tipo_item='producao',
                    criado_via_importacao_estoque=True,
                )
                db.session.add(item)
                db.session.flush()

            estoque_existente = EstoquePecas.query.filter_by(item_id=item.id).first()
            data_entrada = None
            if r.get('data_entrada'):
                data_entrada = datetime.strptime(r['data_entrada'], '%Y-%m-%d').date()
            if not data_entrada:
                data_entrada = datetime.now().date()

            if estoque_existente:
                estoque_existente.quantidade += int(r['quantidade'])
                if r.get('observacao'):
                    estoque_existente.observacao = r['observacao']
                estoque_row = estoque_existente
            else:
                estoque_row = EstoquePecas(
                    item_id=item.id,
                    quantidade=int(r['quantidade']),
                    data_entrada=data_entrada,
                    observacao=r.get('observacao') or None,
                )
                db.session.add(estoque_row)
                db.session.flush()

            movimentacao = MovimentacaoEstoquePecas(
                estoque_pecas_id=estoque_row.id,
                tipo='entrada',
                quantidade=int(r['quantidade']),
                data=data_entrada,
                referencia='IMPORTACAO_EXCEL_ESTOQUE',
                observacao=r.get('observacao') or 'Importação por Excel',
            )
            db.session.add(movimentacao)
            importados += 1

        session['import_estoque_excel_rows'] = selecionados
        db.session.commit()
        flash(f'Importação concluída com sucesso! Linhas importadas: {importados}', 'success')
        return redirect(url_for('estoque_pecas.imprimir_importacao_estoque_excel'))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao importar estoque: {e}', 'danger')
        return redirect(url_for('estoque_pecas.importar_estoque_excel'))


@estoque_pecas.route('/estoque-pecas/importar-excel/imprimir')
def imprimir_importacao_estoque_excel():
    rows = _get_pending_import_rows()
    if not rows:
        flash('Nenhuma importação disponível para impressão.', 'warning')
        return redirect(url_for('estoque_pecas.importar_estoque_excel'))
    return render_template('estoque_pecas/importar_excel_imprimir.html', **_build_import_print_context(rows))

@estoque_pecas.route('/estoque-pecas/entrada', methods=['GET', 'POST'])
def entrada():
    """Rota para registrar entrada de peças no estoque"""
    if request.method == 'POST':
        # Validação de dados
        errors = validate_form_data(request.form, ['item_id', 'quantidade'])
        if errors:
            for error in errors:
                flash(error, 'danger')
            itens = Item.query.all()
            return render_template('estoque_pecas/entrada.html', itens=itens)
        
        item_id = request.form['item_id']
        
        # Validar quantidade
        try:
            quantidade = int(request.form['quantidade'])
            if quantidade <= 0:
                flash('A quantidade deve ser maior que zero', 'danger')
                itens = Item.query.all()
                return render_template('estoque_pecas/entrada.html', itens=itens)
        except ValueError:
            flash('A quantidade deve ser um número inteiro', 'danger')
            itens = Item.query.all()
            return render_template('estoque_pecas/entrada.html', itens=itens)
        
        referencia = request.form.get('referencia', '')
        observacao = request.form.get('observacao', '')
        estante = request.form.get('estante')
        secao = request.form.get('secao')
        linha = request.form.get('linha')
        coluna = request.form.get('coluna')

        def _to_int(v):
            try:
                if v is None or str(v).strip() == '':
                    return None
                return int(v)
            except Exception:
                return None

        estante_i = _to_int(estante)
        secao_i = _to_int(secao)
        linha_i = _to_int(linha)
        coluna_i = _to_int(coluna)

        if estante_i and secao_i and linha_i and coluna_i:
            if coluna_i < 1 or coluna_i > 6:
                flash('Coluna inválida para o mapa (use 1 a 6).', 'warning')
                itens = Item.query.all()
                return render_template('estoque_pecas/entrada.html', itens=itens)
            ocupado = (
                EstoquePecas.query
                .filter(
                    EstoquePecas.estante == estante_i,
                    EstoquePecas.secao == secao_i,
                    EstoquePecas.linha == linha_i,
                    EstoquePecas.coluna == coluna_i,
                )
                .first()
            )
            if ocupado and str(ocupado.item_id) != str(item_id):
                flash('Este endereço já está ocupado por outro item. Escolha outra posição.', 'warning')
                itens = Item.query.all()
                return render_template('estoque_pecas/entrada.html', itens=itens)
        
        # Verificar se já existe um registro para este item
        estoque_existente = EstoquePecas.query.filter_by(item_id=item_id).first()
        
        if estoque_existente:
            # Atualizar estoque existente
            estoque_existente.quantidade += quantidade

            if estante_i:
                estoque_existente.estante = estante_i
            if secao_i:
                estoque_existente.secao = secao_i
            if linha_i:
                estoque_existente.linha = linha_i
            if coluna_i:
                estoque_existente.coluna = coluna_i
            
            # Registrar movimentação
            movimentacao = MovimentacaoEstoquePecas(
                estoque_pecas_id=estoque_existente.id,
                tipo='entrada',
                quantidade=quantidade,
                data=datetime.now().date(),
                referencia=referencia,
                observacao=observacao
            )
            
            db.session.add(movimentacao)
        else:
            # Criar novo registro de estoque
            novo_estoque = EstoquePecas(
                item_id=item_id,
                quantidade=quantidade,
                data_entrada=datetime.now().date(),
                estante=estante_i,
                secao=secao_i,
                linha=linha_i,
                coluna=coluna_i,
                observacao=observacao
            )
            
            db.session.add(novo_estoque)
            db.session.flush()  # Para obter o ID do novo registro
            
            # Registrar movimentação
            movimentacao = MovimentacaoEstoquePecas(
                estoque_pecas_id=novo_estoque.id,
                tipo='entrada',
                quantidade=quantidade,
                data=datetime.now().date(),
                referencia=referencia,
                observacao=observacao
            )
            
            db.session.add(movimentacao)
        
        db.session.commit()
        flash('Entrada de peças registrada com sucesso!', 'success')
        return redirect(url_for('estoque_pecas.index'))
    
    itens = Item.query.all()
    return render_template('estoque_pecas/entrada.html', itens=itens)

@estoque_pecas.route('/estoque-pecas/saida', methods=['GET', 'POST'])
def saida():
    """Rota para registrar saída de peças do estoque"""
    if request.method == 'POST':
        # Validação de dados
        errors = validate_form_data(request.form, ['estoque_id', 'quantidade'])
        if errors:
            for error in errors:
                flash(error, 'danger')
            estoque = EstoquePecas.query.order_by(EstoquePecas.estante, EstoquePecas.secao, EstoquePecas.linha, EstoquePecas.coluna).all()
            return render_template('estoque_pecas/saida.html', estoque=estoque)
        
        estoque_id = request.form['estoque_id']
        
        # Validar quantidade
        try:
            quantidade = int(request.form['quantidade'])
            if quantidade <= 0:
                flash('A quantidade deve ser maior que zero', 'danger')
                estoque = EstoquePecas.query.order_by(EstoquePecas.estante, EstoquePecas.secao, EstoquePecas.linha, EstoquePecas.coluna).all()
                return render_template('estoque_pecas/saida.html', estoque=estoque)
        except ValueError:
            flash('A quantidade deve ser um número inteiro', 'danger')
            estoque = EstoquePecas.query.order_by(EstoquePecas.estante, EstoquePecas.secao, EstoquePecas.linha, EstoquePecas.coluna).all()
            return render_template('estoque_pecas/saida.html', estoque=estoque)
        
        referencia = request.form.get('referencia', '')
        observacao = request.form.get('observacao', '')
        
        # Buscar registro de estoque
        estoque_item = EstoquePecas.query.get_or_404(estoque_id)
        
        # Verificar se há quantidade suficiente
        if estoque_item.quantidade < quantidade:
            flash('Quantidade insuficiente em estoque!', 'danger')
            estoque = EstoquePecas.query.order_by(EstoquePecas.estante, EstoquePecas.secao, EstoquePecas.linha, EstoquePecas.coluna).all()
            return render_template('estoque_pecas/saida.html', estoque=estoque)
        
        # Atualizar estoque
        estoque_item.quantidade -= quantidade
        
        # Registrar movimentação
        movimentacao = MovimentacaoEstoquePecas(
            estoque_pecas_id=estoque_item.id,
            tipo='saida',
            quantidade=quantidade,
            data=datetime.now().date(),
            referencia=referencia,
            observacao=observacao
        )
        
        db.session.add(movimentacao)
        db.session.commit()
        
        flash('Saída de peças registrada com sucesso!', 'success')
        return redirect(url_for('estoque_pecas.index'))
    
    estoque = EstoquePecas.query.order_by(EstoquePecas.estante, EstoquePecas.secao, EstoquePecas.linha, EstoquePecas.coluna).all()
    return render_template('estoque_pecas/saida.html', estoque=estoque)

@estoque_pecas.route('/estoque-pecas/historico/<int:estoque_id>')
def historico(estoque_id):
    """Rota para visualizar o histórico de movimentações de um item do estoque"""
    estoque_item = EstoquePecas.query.get_or_404(estoque_id)
    movimentacoes = MovimentacaoEstoquePecas.query.filter_by(estoque_pecas_id=estoque_id).order_by(MovimentacaoEstoquePecas.data.desc()).all()
    
    return render_template('estoque_pecas/historico.html', estoque=estoque_item, movimentacoes=movimentacoes)


@estoque_pecas.route('/estoque-pecas/mapa')
def mapa():
    estante = request.args.get('estante', '1')
    try:
        estante_i = int(estante) if estante is not None and str(estante).strip() != '' else 1
    except Exception:
        estante_i = 1
    if estante_i < 1 or estante_i > 8:
        estante_i = 1

    def _fetch_mapa_data():
        itens = EstoquePecas.query.order_by(EstoquePecas.id.desc()).all()
        ocup = (
            EstoquePecas.query
            .filter(EstoquePecas.estante == estante_i)
            .all()
        )
        temps = (
            EstoquePecasSlotTemp.query
            .filter(EstoquePecasSlotTemp.estante == estante_i)
            .all()
        )
        return itens, ocup, temps

    try:
        itens_estoque, ocupados, slots_temp = _fetch_mapa_data()
    except ProgrammingError as e:
        msg = str(e or '')
        if 'linha_fim' not in msg:
            raise
        try:
            from migrations.add_estoque_pecas_linha_fim import migrate_postgres, migrate_sqlite
            db_url = current_app.config.get('SQLALCHEMY_DATABASE_URI', '') or ''
            if str(db_url).lower().startswith('postgresql'):
                migrate_postgres()
            else:
                migrate_sqlite()
            db.session.remove()
            itens_estoque, ocupados, slots_temp = _fetch_mapa_data()
        except Exception:
            raise

    # Slots temporários da estante
    slots_temp = (
        EstoquePecasSlotTemp.query
        .filter(EstoquePecasSlotTemp.estante == estante_i)
        .order_by(EstoquePecasSlotTemp.id.desc())
        .all()
    )
    ocupados = (
        EstoquePecas.query
        .filter(EstoquePecas.estante == estante_i)
        .filter(EstoquePecas.slot_temp_id.is_(None))
        .all()
    )

    # Itens que estão em slots temporários desta estante
    slot_ids = [s.id for s in slots_temp]
    itens_em_temp = []
    if slot_ids:
        itens_em_temp = (
            EstoquePecas.query
            .filter(EstoquePecas.slot_temp_id.in_(slot_ids))
            .all()
        )

    def _slot_index(linha_v: int, coluna_v: int) -> int:
        return ((int(linha_v) - 1) * 6) + int(coluna_v)

    def _index_to_cell(idx: int):
        idx_i = int(idx)
        lin = 1 if idx_i <= 6 else 2
        col = idx_i if idx_i <= 6 else (idx_i - 6)
        return lin, col

    ocupados_map = {}
    for e in ocupados:
        for s in _slots_from_entity(e):
            if int(s['estante']) != estante_i:
                continue
            k = (int(s['secao']), int(s['linha']), int(s['coluna']))
            if k not in ocupados_map:
                ocupados_map[k] = []
            ocupados_map[k].append(e)

    # Mapa de slots temporários por célula (para renderização)
    temp_map = {}
    for s in slots_temp:
        for slot in _slots_from_entity(s):
            if int(slot['estante']) != estante_i:
                continue
            temp_map[(int(slot['secao']), int(slot['linha']), int(slot['coluna']))] = s

    # Itens dentro de temporário aparecem como ocupados no endereço do temporário
    slot_by_id = {s.id: s for s in slots_temp}
    for e in itens_em_temp:
        s = slot_by_id.get(e.slot_temp_id)
        if not s:
            continue
        for slot in _slots_from_entity(s):
            if int(slot['estante']) != estante_i:
                continue
            k = (int(slot['secao']), int(slot['linha']), int(slot['coluna']))
            if k not in ocupados_map:
                ocupados_map[k] = []
            ocupados_map[k].append(e)

    return render_template(
        'estoque_pecas/mapa.html',
        estante=estante_i,
        ocupados_map=ocupados_map,
        temp_map=temp_map,
        itens_estoque=itens_estoque,
    )


@estoque_pecas.route('/estoque-pecas/mapa/search')
def mapa_search():
    q = (request.args.get('q') or '').strip()
    if not q:
        return jsonify({'results': []})

    like = f"%{q}%"
    rows = (
        EstoquePecas.query
        .join(Item, EstoquePecas.item_id == Item.id)
        .filter(
            or_(
                Item.codigo_acb.ilike(like),
                Item.nome.ilike(like),
            )
        )
        .filter(EstoquePecas.estante.isnot(None))
        .filter(EstoquePecas.secao.isnot(None))
        .filter(EstoquePecas.linha.isnot(None))
        .filter(EstoquePecas.coluna.isnot(None))
        .order_by(EstoquePecas.estante, EstoquePecas.secao, EstoquePecas.linha, EstoquePecas.coluna)
        .limit(40)
        .all()
    )

    results = []
    for e in rows:
        results.append({
            'estoque_id': e.id,
            'item_id': e.item.id if e.item else None,
            'codigo': (e.item.codigo_acb if e.item else '') or '',
            'nome': (e.item.nome if e.item else '') or '',
            'imagem': (e.item.imagem_path if e.item and e.item.imagem_path else '') or '',
            'pdf': url_for('itens.desenho_pdf_item', item_id=e.item.id) if e.item and e.item.desenho_tecnico else '',
            'estante': e.estante,
            'secao': e.secao,
            'linha': e.linha,
            'coluna': e.coluna,
        })

    return jsonify({'results': results})


@estoque_pecas.route('/estoque-pecas/mapa/definir', methods=['POST'])
def definir_localizacao_mapa():
    estoque_id = request.form.get('estoque_id')
    estante = request.form.get('estante')
    secao = request.form.get('secao')
    linha = request.form.get('linha')
    coluna = request.form.get('coluna')
    linha_fim = request.form.get('linha_fim')
    coluna_fim = request.form.get('coluna_fim')
    slots_json = request.form.get('slots_json')
    permitir_compartilhado = (request.form.get('permitir_compartilhado') or '').strip().lower() in ('1', 'true', 'yes', 'sim', 'on')
    usar_temporario = (request.form.get('usar_temporario') or '').strip().lower() in ('1', 'true', 'yes', 'sim', 'on')
    temporario_nome = (request.form.get('temporario_nome') or '').strip()

    def _to_int(v):
        try:
            if v is None or str(v).strip() == '':
                return None
            return int(v)
        except Exception:
            return None

    estoque_id_i = _to_int(estoque_id)
    estante_i = _to_int(estante)
    secao_i = _to_int(secao)
    linha_i = _to_int(linha)
    coluna_i = _to_int(coluna)
    linha_fim_i = _to_int(linha_fim)
    coluna_fim_i = _to_int(coluna_fim)
    slots_sel = _normalize_slots(_load_slots_json(slots_json, estante_padrao=estante_i), estante_padrao=estante_i)

    if not estoque_id_i and not usar_temporario:
        flash('Selecione um item do estoque (ou marque Temporário).', 'danger')
        return redirect(url_for('estoque_pecas.mapa', estante=estante_i or 1))

    if not (estante_i and secao_i and linha_i and coluna_i):
        flash('Informe Estante, Seção, Linha e Coluna.', 'danger')
        return redirect(url_for('estoque_pecas.mapa', estante=estante_i or 1))

    if estante_i < 1 or estante_i > 8 or secao_i < 1 or secao_i > 4 or linha_i < 1 or linha_i > 2 or coluna_i < 1 or coluna_i > 6:
        flash('Endereço inválido.', 'danger')
        return redirect(url_for('estoque_pecas.mapa', estante=estante_i or 1))

    if not slots_sel:
        slots_sel = _normalize_slots([{'estante': estante_i, 'secao': secao_i, 'linha': linha_i, 'coluna': coluna_i}], estante_padrao=estante_i)

    if linha_fim_i is not None:
        if linha_fim_i < linha_i:
            linha_fim_i = linha_i
        if linha_fim_i < 1 or linha_fim_i > 2:
            flash('Linha final inválida.', 'danger')
            return redirect(url_for('estoque_pecas.mapa', estante=estante_i or 1))

    if coluna_fim_i is not None:
        if coluna_fim_i < coluna_i:
            coluna_fim_i = coluna_i
        if coluna_fim_i < 1 or coluna_fim_i > 6:
            flash('Coluna final inválida.', 'danger')
            return redirect(url_for('estoque_pecas.mapa', estante=estante_i or 1))

    estoque_item = EstoquePecas.query.get(estoque_id_i) if estoque_id_i else None

    ocupacoes = (
        EstoquePecas.query
        .filter(EstoquePecas.estante == estante_i)
        .all()
    )

    slots_sel_set = {(int(s['estante']), int(s['secao']), int(s['linha']), int(s['coluna'])) for s in slots_sel}

    conflita = []
    for o in ocupacoes:
        if estoque_item and o.id == estoque_item.id:
            continue
        o_slots_set = {(int(s['estante']), int(s['secao']), int(s['linha']), int(s['coluna'])) for s in _slots_from_entity(o)}
        if slots_sel_set & o_slots_set:
            conflita.append(o)

    if conflita:
        if not permitir_compartilhado:
            flash('Esta posição já está ocupada por outro item. Marque "Permitir compartilhado" para permitir mais de um item no mesmo slot.', 'warning')
            return redirect(url_for('estoque_pecas.mapa', estante=estante_i))
        # Só permite compartilhar se todas as ocupações já estiverem marcadas como compartilháveis
        if any(not getattr(o, 'permitir_compartilhado', False) for o in conflita):
            flash('Esta posição já tem item(s) que não permitem compartilhamento. Ajuste o(s) item(ns) existente(s) para permitir compartilhamento antes.', 'warning')
            return redirect(url_for('estoque_pecas.mapa', estante=estante_i))

    # Se for temporário, criamos (ou atualizamos) o slot temporário neste endereço.
    slot_temp = None
    if usar_temporario:
        # Só pode haver 1 temporário por célula de início
        slot_temp = (
            EstoquePecasSlotTemp.query
            .filter_by(estante=estante_i, secao=secao_i, linha=linha_i, coluna=coluna_i)
            .first()
        )
        if not slot_temp:
            slot_temp = EstoquePecasSlotTemp(
                estante=estante_i,
                secao=secao_i,
                linha=linha_i,
                coluna=coluna_i,
            )
            db.session.add(slot_temp)
            db.session.flush()

        slot_temp.nome = temporario_nome or slot_temp.nome
        slot_temp.linha_fim = linha_fim_i
        slot_temp.coluna_fim = coluna_fim_i
        slot_temp.estante = int(slots_sel[0]['estante'])
        slot_temp.secao = int(slots_sel[0]['secao'])
        slot_temp.linha = int(slots_sel[0]['linha'])
        slot_temp.coluna = int(slots_sel[0]['coluna'])
        slot_temp.slots_json = _dump_slots_json(slots_sel)
        slot_temp.permitir_compartilhado = True if permitir_compartilhado else False

    # Se selecionou item, atualiza local.
    if estoque_item:
        if slot_temp:
            # Endereço passa a ser regido pelo temporário
            estoque_item.slot_temp_id = slot_temp.id
            estoque_item.estante = None
            estoque_item.secao = None
            estoque_item.linha = None
            estoque_item.coluna = None
            estoque_item.linha_fim = None
            estoque_item.coluna_fim = None
            estoque_item.slots_json = None
            estoque_item.permitir_compartilhado = permitir_compartilhado
        else:
            estoque_item.slot_temp_id = None
            estoque_item.estante = int(slots_sel[0]['estante'])
            estoque_item.secao = int(slots_sel[0]['secao'])
            estoque_item.linha = int(slots_sel[0]['linha'])
            estoque_item.coluna = int(slots_sel[0]['coluna'])
            estoque_item.linha_fim = linha_fim_i
            estoque_item.coluna_fim = coluna_fim_i
            estoque_item.slots_json = _dump_slots_json(slots_sel)
            estoque_item.permitir_compartilhado = permitir_compartilhado

    db.session.commit()
    flash('Localização atualizada com sucesso!', 'success')
    return redirect(url_for('estoque_pecas.mapa', estante=estante_i))


@estoque_pecas.route('/estoque-pecas/mapa/remover', methods=['POST'])
def remover_localizacao_mapa():
    estoque_id = request.form.get('estoque_id')
    estante = request.form.get('estante')

    def _to_int(v):
        try:
            if v is None or str(v).strip() == '':
                return None
            return int(v)
        except Exception:
            return None

    estoque_id_i = _to_int(estoque_id)
    estante_i = _to_int(estante) or 1
    if not estoque_id_i:
        flash('Selecione um item do estoque.', 'danger')
        return redirect(url_for('estoque_pecas.mapa', estante=estante_i))

    estoque_item = EstoquePecas.query.get_or_404(estoque_id_i)
    slot_temp_id = getattr(estoque_item, 'slot_temp_id', None)

    estoque_item.slot_temp_id = None
    estoque_item.estante = None
    estoque_item.secao = None
    estoque_item.linha = None
    estoque_item.coluna = None
    estoque_item.linha_fim = None
    estoque_item.coluna_fim = None
    estoque_item.slots_json = None
    db.session.commit()

    # Se estava em temporário, remover o temporário quando ficar vazio
    if slot_temp_id:
        rest = EstoquePecas.query.filter_by(slot_temp_id=slot_temp_id).first()
        if not rest:
            slot = EstoquePecasSlotTemp.query.get(slot_temp_id)
            if slot:
                db.session.delete(slot)
                db.session.commit()

    flash('Localização removida com sucesso!', 'success')
    return redirect(url_for('estoque_pecas.mapa', estante=estante_i))


@estoque_pecas.route('/estoque-pecas/mapa/temp/remover', methods=['POST'])
def remover_slot_temporario_mapa():
    estante = request.form.get('estante')
    secao = request.form.get('secao')
    linha = request.form.get('linha')
    coluna = request.form.get('coluna')

    def _to_int(v):
        try:
            if v is None or str(v).strip() == '':
                return None
            return int(v)
        except Exception:
            return None

    estante_i = _to_int(estante) or 1
    secao_i = _to_int(secao)
    linha_i = _to_int(linha)
    coluna_i = _to_int(coluna)

    if not (secao_i and linha_i and coluna_i):
        flash('Endereço inválido para remover temporário.', 'danger')
        return redirect(url_for('estoque_pecas.mapa', estante=estante_i))

    slot = (
        EstoquePecasSlotTemp.query
        .filter_by(estante=estante_i, secao=secao_i, linha=linha_i, coluna=coluna_i)
        .first()
    )
    if not slot:
        flash('Slot temporário não encontrado.', 'warning')
        return redirect(url_for('estoque_pecas.mapa', estante=estante_i))

    # Só remove se estiver vazio
    existe = EstoquePecas.query.filter_by(slot_temp_id=slot.id).first()
    if existe:
        flash('Este temporário ainda possui itens. Remova os itens primeiro.', 'warning')
        return redirect(url_for('estoque_pecas.mapa', estante=estante_i))

    db.session.delete(slot)
    db.session.commit()
    flash('Slot temporário removido.', 'success')
    return redirect(url_for('estoque_pecas.mapa', estante=estante_i))

@estoque_pecas.route('/estoque-pecas/atualizar-localizacao/<int:estoque_id>', methods=['POST'])
def atualizar_localizacao(estoque_id):
    """Rota para atualizar a localização (prateleira e posição) de um item no estoque"""
    estoque_item = EstoquePecas.query.get_or_404(estoque_id)
    
    if request.method == 'POST':
        estante = request.form.get('estante')
        secao = request.form.get('secao')
        slot = request.form.get('slot')
        slot_fim = request.form.get('slot_fim')
        permitir_compartilhado = (request.form.get('permitir_compartilhado') or '').strip().lower() in ('1', 'true', 'yes', 'sim', 'on')

        def _to_int(v):
            try:
                if v is None or str(v).strip() == '':
                    return None
                return int(v)
            except Exception:
                return None
        
        estoque_item.prateleira = None
        estoque_item.posicao = None

        estante_i = _to_int(estante)
        secao_i = _to_int(secao)
        slot_i = _to_int(slot)
        slot_fim_i = _to_int(slot_fim)

        # Converter slot para linha e coluna
        if slot_i:
            if slot_i < 1 or slot_i > 24:
                flash('Slot inválido (use 1 a 24).', 'warning')
                return redirect(url_for('estoque_pecas.index'))
            
            # Calcular linha e coluna a partir do slot
            if slot_i <= 12:
                linha_i = 1
                coluna_i = slot_i
            else:
                linha_i = 2
                coluna_i = slot_i - 12
        else:
            linha_i = None
            coluna_i = None

        # Converter slot_fim para linha_fim e coluna_fim
        if slot_fim_i:
            if slot_fim_i < 1 or slot_fim_i > 24:
                flash('Slot final inválido (use 1 a 24).', 'warning')
                return redirect(url_for('estoque_pecas.index'))
            
            if slot_fim_i <= 12:
                linha_fim_i = 1
                coluna_fim_i = slot_fim_i
            else:
                linha_fim_i = 2
                coluna_fim_i = slot_fim_i - 12
        else:
            linha_fim_i = None
            coluna_fim_i = None

        if estante_i and secao_i and linha_i and coluna_i:
            # Validação de coluna já foi feita acima
            if coluna_fim_i is not None:
                if coluna_fim_i < coluna_i:
                    coluna_fim_i = coluna_i
                if coluna_fim_i < 1 or coluna_fim_i > 12:
                    flash('Coluna final inválida para o mapa (use 1 a 12).', 'warning')
                    return redirect(url_for('estoque_pecas.index'))

            # Para múltiplas linhas, verificar cada linha separadamente
            if linha_fim_i and linha_fim_i != linha_i:
                # Verificar conflitos em todas as linhas entre linha_i e linha_fim_i
                for check_linha in range(min(linha_i, linha_fim_i), max(linha_i, linha_fim_i) + 1):
                    col_ini_check = coluna_i if check_linha == linha_i else 1
                    col_fim_check = coluna_fim_i if check_linha == linha_fim_i else 12
                    
                    ocupacoes_linha = (
                        EstoquePecas.query
                        .filter(
                            EstoquePecas.estante == estante_i,
                            EstoquePecas.secao == secao_i,
                            EstoquePecas.linha == check_linha,
                            EstoquePecas.id != estoque_item.id
                        )
                        .all()
                    )

                    def _ranges_overlap(a1, a2, b1, b2):
                        return max(a1, b1) <= min(a2, b2)

                    conflita = []
                    for o in ocupacoes_linha:
                        o_ini = int(o.coluna) if o.coluna else None
                        if not o_ini:
                            continue
                        o_fim = int(o.coluna_fim) if getattr(o, 'coluna_fim', None) else o_ini
                        o_fim = max(o_ini, min(o_fim, 12))
                        if _ranges_overlap(col_ini_check, col_fim_check, o_ini, o_fim):
                            conflita.append(o)

                    if conflita:
                        if not permitir_compartilhado:
                            flash('Esta posição já está ocupada por outro item. Marque "Permitir compartilhado" para permitir mais de um item no mesmo slot.', 'warning')
                            return redirect(url_for('estoque_pecas.index'))
                        if any(not getattr(o, 'permitir_compartilhado', False) for o in conflita):
                            flash('Esta posição já tem item(s) que não permitem compartilhamento. Ajuste o(s) item(ns) existente(s) para permitir compartilhamento antes.', 'warning')
                            return redirect(url_for('estoque_pecas.index'))
            else:
                # Verificação normal para mesma linha
                col_fim_check = coluna_fim_i if coluna_fim_i is not None else coluna_i
                col_ini_check = coluna_i
                ocupacoes = (
                    EstoquePecas.query
                    .filter(
                        EstoquePecas.estante == estante_i,
                        EstoquePecas.secao == secao_i,
                        EstoquePecas.linha == linha_i,
                        EstoquePecas.id != estoque_item.id
                    )
                    .all()
                )

                def _ranges_overlap(a1, a2, b1, b2):
                    return max(a1, b1) <= min(a2, b2)

                conflita = []
                for o in ocupacoes:
                    o_ini = int(o.coluna) if o.coluna else None
                    if not o_ini:
                        continue
                    o_fim = int(o.coluna_fim) if getattr(o, 'coluna_fim', None) else o_ini
                    o_fim = max(o_ini, min(o_fim, 12))
                    if _ranges_overlap(col_ini_check, col_fim_check, o_ini, o_fim):
                        conflita.append(o)

                if conflita:
                    if not permitir_compartilhado:
                        flash('Esta posição já está ocupada por outro item. Marque "Permitir compartilhado" para permitir mais de um item no mesmo slot.', 'warning')
                        return redirect(url_for('estoque_pecas.index'))
                    if any(not getattr(o, 'permitir_compartilhado', False) for o in conflita):
                        flash('Esta posição já tem item(s) que não permitem compartilhamento. Ajuste o(s) item(ns) existente(s) para permitir compartilhamento antes.', 'warning')
                        return redirect(url_for('estoque_pecas.index'))

        estoque_item.estante = estante_i
        estoque_item.secao = secao_i
        estoque_item.linha = linha_i
        estoque_item.coluna = coluna_i
        estoque_item.linha_fim = linha_fim_i
        estoque_item.coluna_fim = coluna_fim_i
        # Posição é determinada automaticamente pela coluna (1-6=frente, 7-12=trás)
        if coluna_i and coluna_i <= 6:
            estoque_item.posicao = 'frente'
        elif coluna_i and coluna_i > 6:
            estoque_item.posicao = 'tras'
        else:
            estoque_item.posicao = None
        estoque_item.permitir_compartilhado = permitir_compartilhado
        
        db.session.commit()
        flash('Localização atualizada com sucesso!', 'success')
    
    return redirect(url_for('estoque_pecas.index'))

@estoque_pecas.route('/estoque-pecas/movimentacao-rapida/<int:estoque_id>/<string:tipo>', methods=['POST'])
def movimentacao_rapida(estoque_id, tipo):
    """Rota para movimentação rápida (entrada ou saída) de itens no estoque"""
    estoque_item = EstoquePecas.query.get_or_404(estoque_id)
    
    if request.method == 'POST':
        try:
            quantidade = int(request.form.get('quantidade', 1))
            if quantidade <= 0:
                flash('A quantidade deve ser maior que zero', 'danger')
                return redirect(url_for('estoque_pecas.index'))
        except ValueError:
            flash('A quantidade deve ser um número inteiro', 'danger')
            return redirect(url_for('estoque_pecas.index'))
        
        referencia = request.form.get('referencia', '')
        observacao = request.form.get('observacao', 'Movimentação rápida')
        
        if tipo == 'entrada':
            estoque_item.quantidade += quantidade
        elif tipo == 'saida':
            if estoque_item.quantidade < quantidade:
                flash('Quantidade insuficiente em estoque!', 'danger')
                return redirect(url_for('estoque_pecas.index'))
            estoque_item.quantidade -= quantidade
        else:
            flash('Tipo de movimentação inválido!', 'danger')
            return redirect(url_for('estoque_pecas.index'))
        
        # Registrar movimentação
        movimentacao = MovimentacaoEstoquePecas(
            estoque_pecas_id=estoque_item.id,
            tipo=tipo,
            quantidade=quantidade,
            data=datetime.now().date(),
            referencia=referencia,
            observacao=observacao
        )
        
        db.session.add(movimentacao)
        db.session.commit()
        
        flash(f'Movimentação de {quantidade} item(ns) registrada com sucesso!', 'success')
    
    return redirect(url_for('estoque_pecas.index'))


@estoque_pecas.route('/estoque/valores')
def valores_estoque():
    """Página de estoque com valores - requer permissão especial"""
    acesso_negado = _require_acesso_valores()
    if acesso_negado:
        return acesso_negado
    
    # Buscar todos os itens em estoque com suas quantidades
    # Usando sintaxe compatível com PostgreSQL
    from sqlalchemy import func, case, text
    
    query = db.session.query(
        Item.id,
        Item.codigo_acb,
        Item.nome,
        Item.imagem,
        Item.valor_item,
        func.coalesce(func.sum(EstoquePecas.quantidade), 0).label('quantidade_total')
    ).outerjoin(
        EstoquePecas, Item.id == EstoquePecas.item_id
    ).group_by(
        Item.id, Item.codigo_acb, Item.nome, Item.imagem, Item.valor_item
    ).having(
        func.sum(EstoquePecas.quantidade) > 0
    ).order_by(
        Item.codigo_acb
    ).all()
    
    # Buscar localizações separadamente para evitar problemas com group_concat
    item_ids = [item.id for item in query]
    locais_map = {}
    
    if item_ids:
        locais_query = db.session.query(
            EstoquePecas.item_id,
            func.concat(
                'E', EstoquePecas.estante, 
                'S', EstoquePecas.secao, 
                'L', EstoquePecas.linha, 
                'C', EstoquePecas.coluna
            ).label('local')
        ).filter(
            EstoquePecas.item_id.in_(item_ids),
            EstoquePecas.estante.isnot(None)
        ).distinct().all()
        
        for loc in locais_query:
            if loc.item_id not in locais_map:
                locais_map[loc.item_id] = []
            locais_map[loc.item_id].append(loc.local)
    
    # Calcular totais
    total_pecas = 0
    total_valor = 0.0
    
    itens_estoque = []
    for item in query:
        quantidade = int(item.quantidade_total)
        valor_unitario = float(item.valor_item or 0)
        valor_total = quantidade * valor_unitario
        
        total_pecas += quantidade
        total_valor += valor_total
        
        # Formatar localização usando o mapa de locais
        locais_lista = locais_map.get(item.id, [])
        if locais_lista:
            # Remover duplicados e ordenar
            locais_unicos = sorted(set(locais_lista))
            locais = ', '.join(locais_unicos[:3])  # Mostrar até 3 locais
            if len(locais_unicos) > 3:
                locais += f' (+{len(locais_unicos) - 3})'
        else:
            locais = ''
        
        itens_estoque.append({
            'id': item.id,
            'codigo_acb': item.codigo_acb,
            'nome': item.nome,
            'imagem_url': get_file_url(item.imagem) if item.imagem else None,
            'quantidade': quantidade,
            'locais': locais,
            'valor_unitario': valor_unitario,
            'valor_total': valor_total
        })
    
    return render_template('estoque_pecas/valores.html', 
                         itens=itens_estoque,
                         total_pecas=total_pecas,
                         total_valor=total_valor)
